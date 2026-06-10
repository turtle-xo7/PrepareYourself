"""Excel export for superadmin.

Split from the original monolithic core/views.py.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.conf import settings
from django.db import transaction
from datetime import datetime
import json
import uuid

from ..models import Board, Subject, Class, Question, UserProfile, UserProgress
from ..services import ai as ai_svc
from .base import (
    logger, _L, _upload_error, ALLOWED_IMAGE_EXTS, ALLOWED_DOC_EXTS,
    CURRENT_YEAR, YEARS,
    admin_required, superadmin_required, premium_required,
    _is_exam_staff, _notify_all_students,
)

@superadmin_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from ..models import UserProgress, TeacherFeedback, StudyNote, Contest, ContestSubmission, NoteBookmark, ExamPaper, ExamAttempt, NoteRequest, WrittenSolveSubmission

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')

    def style_header(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(len(header) + 5, 15)

    ws1 = wb.active
    ws1.title = 'Users'
    headers = ['Username', 'Email', 'Role', 'Plan', 'Is Superadmin', 'Joined']
    style_header(ws1, headers)
    for profile in UserProfile.objects.select_related('user').all():
        ws1.append([
            profile.user.username,
            profile.user.email,
            profile.role,
            profile.plan,
            'Yes' if profile.is_superadmin else 'No',
            profile.user.date_joined.strftime('%d-%m-%Y %H:%M'),
        ])

    ws2 = wb.create_sheet('Questions')
    headers2 = ['Board', 'Subject', 'Class', 'Year', 'Chapter', 'Question', 'Type', 'Difficulty', 'Correct Option', 'Answer Hint']
    style_header(ws2, headers2)
    for q in Question.objects.select_related('board', 'subject', 'class_obj').filter(is_active=True):
        ws2.append([
            q.board.name, q.subject.name, q.class_obj.name, q.year,
            q.chapter, q.question_text, q.question_type, q.difficulty,
            q.correct_option, q.answer_hint,
        ])

    ws3 = wb.create_sheet('Progress')
    headers3 = ['Student', 'Question', 'Subject', 'Correct', 'Answered At']
    style_header(ws3, headers3)
    for p in UserProgress.objects.select_related('user', 'question', 'question__subject').all():
        ws3.append([
            p.user.username, p.question.question_text[:50],
            p.question.subject.name, 'Yes' if p.is_correct else 'No',
            p.answered_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws4 = wb.create_sheet('Study Notes')
    headers4 = ['Title', 'Subject', 'Class', 'Chapter', 'Created By', 'Created At']
    style_header(ws4, headers4)
    for note in StudyNote.objects.select_related('subject', 'class_obj', 'created_by').all():
        ws4.append([
            note.title, note.subject.name, note.class_obj.name,
            note.chapter, note.created_by.username,
            note.created_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws5 = wb.create_sheet('Contests')
    headers5 = ['Title', 'Subject', 'Class', 'Created By', 'Duration', 'Start', 'End', 'Submissions']
    style_header(ws5, headers5)
    for c in Contest.objects.select_related('subject', 'class_obj', 'created_by').all():
        ws5.append([
            c.title, c.subject.name, c.class_obj.name, c.created_by.username,
            f"{c.duration_minutes} min",
            c.start_time.strftime('%d-%m-%Y %H:%M'),
            c.end_time.strftime('%d-%m-%Y %H:%M'),
            c.submissions.filter(is_submitted=True).count(),
        ])

    ws6 = wb.create_sheet('Contest Results')
    headers6 = ['Contest', 'Student', 'Total Marks', 'Duration (s)', 'Submitted At']
    style_header(ws6, headers6)
    for sub in ContestSubmission.objects.select_related('contest', 'student').filter(is_submitted=True):
        ws6.append([
            sub.contest.title, sub.student.username, sub.total_marks,
            sub.duration_taken,
            sub.submitted_at.strftime('%d-%m-%Y %H:%M') if sub.submitted_at else '',
        ])

    ws7 = wb.create_sheet('Teacher Feedback')
    headers7 = ['Teacher', 'Student', 'Comment', 'Is Read', 'Created At']
    style_header(ws7, headers7)
    for fb in TeacherFeedback.objects.select_related('teacher', 'student').all():
        ws7.append([
            fb.teacher.username, fb.student.username, fb.comment,
            'Yes' if fb.is_read else 'No',
            fb.created_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws8 = wb.create_sheet('Bookmarks')
    headers8 = ['Student', 'Note Title', 'Bookmarked At']
    style_header(ws8, headers8)
    for bm in NoteBookmark.objects.select_related('user', 'note').all():
        ws8.append([
            bm.user.username, bm.note.title,
            bm.created_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws9 = wb.create_sheet('Exam Papers')
    headers9 = ['Title', 'Subject', 'Class', 'Board', 'Year', 'Created By', 'MCQ Count', 'CQ Count', 'Total Attempts', 'Graded', 'Created At']
    style_header(ws9, headers9)
    for paper in ExamPaper.objects.select_related('subject', 'class_obj', 'board', 'created_by').all():
        total_attempts = paper.attempts.count()
        graded = paper.attempts.filter(status='GRADED').count()
        ws9.append([
            paper.title,
            paper.subject.name,
            paper.class_obj.name,
            paper.board.name if paper.board else '',
            paper.year or '',
            paper.created_by.username,
            paper.mcqs.count(),
            paper.cqs.count(),
            total_attempts,
            graded,
            paper.created_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws10 = wb.create_sheet('Exam Results')
    headers10 = ['Student', 'Exam Paper', 'Subject', 'Status', 'MCQ Score', 'CQ Score', 'Total Score', 'Grade', 'Started At', 'CQ Submitted At']
    style_header(ws10, headers10)
    for attempt in ExamAttempt.objects.select_related('student', 'exam_paper', 'exam_paper__subject').all():
        ws10.append([
            attempt.student.username,
            attempt.exam_paper.title,
            attempt.exam_paper.subject.name,
            attempt.get_status_display(),
            attempt.mcq_score,
            attempt.cq_score if attempt.cq_score is not None else '',
            attempt.total_score if attempt.total_score is not None else '',
            attempt.grade or '',
            attempt.started_at.strftime('%d-%m-%Y %H:%M'),
            attempt.cq_submitted_at.strftime('%d-%m-%Y %H:%M') if attempt.cq_submitted_at else '',
        ])

    ws11 = wb.create_sheet('Note Requests')
    headers11 = ['Student', 'Subject', 'Topic', 'Details', 'Status', 'Requested At', 'Fulfilled At', 'Fulfilled By', 'Fulfilled Note']
    style_header(ws11, headers11)
    for nr in NoteRequest.objects.select_related('student', 'subject', 'fulfilled_by', 'fulfilled_note').all():
        ws11.append([
            nr.student.username,
            nr.subject.name if nr.subject else '',
            nr.topic,
            nr.details,
            nr.status,
            nr.created_at.strftime('%d-%m-%Y %H:%M'),
            nr.fulfilled_at.strftime('%d-%m-%Y %H:%M') if nr.fulfilled_at else '',
            nr.fulfilled_by.username if nr.fulfilled_by else '',
            nr.fulfilled_note.title if nr.fulfilled_note else '',
        ])

    ws12 = wb.create_sheet('Written CQ Submissions')
    headers12 = ['Student', 'Question', 'Subject', 'Chapter', 'Board', 'Year', 'Submitted At']
    style_header(ws12, headers12)
    for ws in WrittenSolveSubmission.objects.select_related('student', 'question', 'question__subject', 'question__board').all():
        ws12.append([
            ws.student.username,
            ws.question.question_text[:60],
            ws.question.subject.name,
            ws.question.chapter or '',
            ws.question.board.name,
            ws.question.year,
            ws.submitted_at.strftime('%d-%m-%Y %H:%M'),
        ])

    from django.utils import timezone
    filename = f"PrepareYourself_Data_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -------- EXAM MODE VIEWS --------

