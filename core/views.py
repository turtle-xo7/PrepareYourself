from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.conf import settings
from django.db import transaction
from .models import Board, Subject, Class, Question, UserProfile, UserProgress
from .services import ai as ai_svc
from datetime import datetime
import json
import logging
import uuid

logger = logging.getLogger('core')

CURRENT_YEAR = datetime.now().year
YEARS = list(range(CURRENT_YEAR, CURRENT_YEAR - 6, -1))


def _L(request, en, bn):
    """Return English or Bangla based on the active language."""
    return en if getattr(request, 'LANG', 'bn') == 'en' else bn


# -------- UPLOAD VALIDATION --------

ALLOWED_IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}
ALLOWED_DOC_EXTS = ALLOWED_IMAGE_EXTS | {'.pdf'}


def _upload_error(uploaded, kind='image', max_mb=10):
    """Validate an uploaded file's extension and size.

    Returns None when acceptable, otherwise an (english, bangla) error pair.
    Pass kind='doc' to additionally allow PDFs.
    """
    if uploaded is None:
        return None
    import os
    allowed = ALLOWED_DOC_EXTS if kind == 'doc' else ALLOWED_IMAGE_EXTS
    ext = os.path.splitext(uploaded.name)[1].lower()
    if ext not in allowed:
        kinds = 'image/PDF' if kind == 'doc' else 'image'
        return (f'"{uploaded.name}" is not an allowed {kinds} file.',
                f'"{uploaded.name}" অনুমোদিত {kinds} ফাইল নয়।')
    if uploaded.size > max_mb * 1024 * 1024:
        return (f'"{uploaded.name}" is too large (max {max_mb} MB).',
                f'"{uploaded.name}" খুব বড় (সর্বোচ্চ {max_mb} MB)।')
    return None


# -------- DECORATORS --------

def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        try:
            profile = request.user.profile
            if profile.role != 'ADMIN':
                messages.error(request, _L(request, 'Only Teachers/Tutors/Institutions can access this page.', 'শুধু Teacher/Tutor/Institution এই page access করতে পারবে।'))
                return redirect('home')
            if not profile.is_approved and not profile.is_superadmin:
                return redirect('teacher_pending')
        except UserProfile.DoesNotExist:
            return redirect('home')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def superadmin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        try:
            if not request.user.profile.is_superadmin:
                return redirect('home')
        except:
            return redirect('home')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


def premium_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        try:
            profile = request.user.profile
            if profile.role == 'ADMIN' or profile.is_superadmin:
                return view_func(request, *args, **kwargs)
            if not profile.is_premium:
                messages.error(request, _L(request, 'This feature is for Premium users only.', 'এই feature শুধু Premium users এর জন্য।'))
                return redirect('pricing')
        except UserProfile.DoesNotExist:
            return redirect('pricing')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


# -------- AUTH VIEWS --------

def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Email দিয়ে login support
        if '@' in username:
            try:
                from django.contrib.auth.models import User
                user_obj = User.objects.get(email=username)
                username = user_obj.username
            except User.DoesNotExist:
                pass

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, _L(request, 'Incorrect username/email or password.', 'Username/Email বা Password ভুল।'))
    return render(request, 'core/login.html')


def signup_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        role = request.POST.get('role', 'STUDENT')
        plan = request.POST.get('plan', 'FREE')
        admin_code = request.POST.get('admin_code', '')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already taken!')
            return redirect('login')

        for doc in (request.FILES.get('nid_document'), request.FILES.get('qualification_document')):
            err = _upload_error(doc, kind='doc', max_mb=10)
            if err:
                messages.error(request, _L(request, *err))
                return redirect('login')

        is_superadmin = False
        if admin_code == 'PY2026ADMIN':
            role = 'ADMIN'
            is_superadmin = True

        is_approved = True
        if role == 'ADMIN' and not is_superadmin:
            is_approved = False

        # User + profile must be created together — an orphan User without a
        # UserProfile breaks every profile-dependent view after login.
        with transaction.atomic():
            user = User.objects.create_user(username=username, email=email, password=password)
            profile = UserProfile.objects.create(
                user=user,
                role=role,
                plan=plan,
                is_superadmin=is_superadmin,
                is_approved=is_approved,
            )

            if role == 'ADMIN' and not is_superadmin:
                profile.teacher_bio = request.POST.get('teacher_bio', '')
                profile.subject_expertise = request.POST.get('subject_expertise', '')
                if request.FILES.get('nid_document'):
                    profile.nid_document = request.FILES['nid_document']
                if request.FILES.get('qualification_document'):
                    profile.qualification_document = request.FILES['qualification_document']
                profile.save()

        logger.info('New signup: %s (role=%s, plan=%s)', username, role, plan)
        login(request, user)

        if role == 'ADMIN' and not is_superadmin:
            return redirect('teacher_pending')

        if plan != 'FREE':
            return redirect(f'/checkout/?plan={plan}')

        # New students go through onboarding first
        return redirect('onboarding')

    return redirect('login')


@login_required
def onboarding(request):
    """One-time setup: capture Board + Class + Subjects so content is personalized."""
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        return redirect('home')

    # Teachers/admins don't onboard as students
    if profile.role == 'ADMIN' or profile.is_superadmin:
        return redirect('home')

    if request.method == 'POST':
        board_id = request.POST.get('board')
        class_id = request.POST.get('class')
        subject_ids = request.POST.getlist('subjects')
        goal = request.POST.get('goal', '').strip()

        if board_id:
            profile.board_id = board_id
        if class_id:
            profile.class_obj_id = class_id
        profile.exam_goal = goal[:120]
        profile.onboarded = True
        profile.save()
        if subject_ids:
            profile.study_subjects.set(subject_ids)

        messages.success(request, 'Your study space is ready! 🎉' if getattr(request, 'LANG', 'bn') == 'en' else 'তোমার স্টাডি স্পেস তৈরি! 🎉')
        return redirect('dashboard' if profile.is_premium else 'question_bank')

    return render(request, 'core/onboarding.html', {
        'boards': Board.objects.filter(is_active=True),
        'classes': Class.objects.all(),
        'subjects': Subject.objects.filter(is_active=True),
    })


def logout_view(request):
    logout(request)
    return redirect('login')


def toggle_language(request):
    current = getattr(request, 'LANG', 'bn')
    new_lang = 'en' if current == 'bn' else 'bn'
    if request.user.is_authenticated:
        try:
            profile = request.user.profile
            profile.preferred_language = new_lang
            profile.save(update_fields=['preferred_language'])
        except Exception:
            pass
    request.session['preferred_language'] = new_lang
    return redirect(request.META.get('HTTP_REFERER', '/'))


PLAN_PRICES = {'FREE': 0, 'BASIC': 99, 'PREMIUM': 199}
PLAN_LABELS = {'FREE': 'Free', 'BASIC': 'Basic', 'PREMIUM': 'Premium'}


def _activate_plan(profile, plan):
    """Single source of truth for granting a paid plan (30 days; extends if still active)."""
    from django.utils import timezone
    from datetime import timedelta
    now = timezone.now()
    base = profile.plan_expires_at if (profile.plan_expires_at and profile.plan_expires_at > now) else now
    profile.plan = plan
    profile.plan_expires_at = base + timedelta(days=30)
    profile.save(update_fields=['plan', 'plan_expires_at'])


@login_required
def checkout(request):
    from .models import Payment
    if request.method == 'POST':
        plan = (request.POST.get('plan') or 'BASIC').upper()
        if plan not in ('BASIC', 'PREMIUM'):
            plan = 'BASIC'
        payment = Payment.objects.create(
            user=request.user,
            plan=plan,
            amount=PLAN_PRICES[plan],
            method=request.POST.get('method', ''),
            tran_id=uuid.uuid4().hex,
            gateway='simulation',
            status='PENDING',
        )
        # Plan is NOT activated here. Activation only happens after a verified
        # payment (payment_process). To go live, replace this redirect with the
        # gateway "initiate" call and redirect to the returned checkout URL.
        return redirect('payment_simulate', tran_id=payment.tran_id)

    plan = request.GET.get('plan') or request.user.profile.plan or 'BASIC'
    plan = plan.upper()
    if plan not in PLAN_PRICES or plan == 'FREE':
        plan = 'BASIC'
    return render(request, 'core/checkout.html', {
        'plan': plan,
        'plan_label': PLAN_LABELS.get(plan, plan.title()),
        'price': PLAN_PRICES.get(plan, 99),
    })


@login_required
def payment_simulate(request, tran_id):
    """Dev-only fake gateway. Disabled when DEBUG is False, so production cannot
    grant free plans before a real gateway is integrated."""
    from .models import Payment
    if not settings.DEBUG:
        messages.error(request, _L(request, 'Online payment is not available yet.', 'অনলাইন পেমেন্ট এখনো চালু হয়নি।'))
        return redirect('pricing')
    payment = get_object_or_404(Payment, tran_id=tran_id, user=request.user)
    if payment.status == 'COMPLETED':
        return redirect('payment_success', tran_id=tran_id)
    if payment.status != 'PENDING':
        return redirect('pricing')
    return render(request, 'core/payment_simulate.html', {'payment': payment})


@login_required
def payment_process(request, tran_id):
    """Completes a simulated payment. In a real gateway this becomes the
    server-side validation callback: verify val_id + amount with the gateway
    BEFORE calling _activate_plan."""
    from .models import Payment
    if request.method != 'POST' or not settings.DEBUG:
        return redirect('pricing')
    payment = get_object_or_404(Payment, tran_id=tran_id, user=request.user)
    if payment.status != 'PENDING':
        return redirect('payment_success', tran_id=tran_id)
    if request.POST.get('result') == 'success':
        # Marking the payment COMPLETED and granting the plan must succeed or
        # fail together — a COMPLETED payment without an active plan (or vice
        # versa) cannot be reconciled later.
        with transaction.atomic():
            payment.status = 'COMPLETED'
            payment.val_id = 'SIM-' + payment.tran_id[:12]
            payment.save(update_fields=['status', 'val_id', 'updated_at'])
            _activate_plan(request.user.profile, payment.plan)
        logger.info('Payment completed: tran_id=%s user=%s plan=%s amount=%s',
                    payment.tran_id, request.user.username, payment.plan, payment.amount)
        return redirect('payment_success', tran_id=tran_id)
    payment.status = 'FAILED'
    payment.save(update_fields=['status', 'updated_at'])
    logger.warning('Payment failed: tran_id=%s user=%s plan=%s',
                   payment.tran_id, request.user.username, payment.plan)
    return redirect('payment_failed')


@login_required
def payment_success(request, tran_id):
    from .models import Payment
    payment = get_object_or_404(Payment, tran_id=tran_id, user=request.user)
    profile = request.user.profile
    needs_onboarding = (not profile.onboarded and profile.role != 'ADMIN'
                        and not profile.is_superadmin)
    return render(request, 'core/payment_success.html', {
        'payment': payment,
        'needs_onboarding': needs_onboarding,
    })


def payment_failed(request):
    return render(request, 'core/payment_failed.html', {
        'error_message': _L(request,
                            'The payment was cancelled or could not be completed.',
                            'পেমেন্ট বাতিল হয়েছে বা সম্পন্ন করা যায়নি।'),
    })


# -------- FRONTEND VIEWS --------

def home(request):
    from .models import Board
    boards = Board.objects.filter(is_active=True)
    return render(request, 'core/home.html', {'boards': boards})

def pricing(request):
    return render(request, 'core/pricing.html')


def _search_scope(request):
    """Return (board_id, class_id) to scope search to, from the user's profile."""
    if request.user.is_authenticated:
        try:
            p = request.user.profile
            return (getattr(p, 'board_id', None), getattr(p, 'class_obj_id', None))
        except UserProfile.DoesNotExist:
            pass
    return (None, None)


def _run_search(query, board_id=None, class_id=None, scoped=True, limit=6):
    """Search questions, notes, exam papers. Returns dict of grouped lists."""
    from .models import StudyNote, ExamPaper
    from django.db.models import Q
    q = (query or '').strip()
    if not q:
        return {'questions': [], 'notes': [], 'papers': []}

    questions = Question.objects.select_related('subject', 'board', 'class_obj').filter(
        is_active=True
    ).filter(
        Q(question_text__icontains=q) | Q(chapter__icontains=q) | Q(subject__name__icontains=q)
    )
    notes = StudyNote.objects.select_related('subject', 'class_obj').filter(
        is_active=True
    ).filter(
        Q(title__icontains=q) | Q(chapter__icontains=q) | Q(content__icontains=q)
    )
    papers = ExamPaper.objects.select_related('subject', 'class_obj', 'board').filter(
        is_active=True
    ).filter(
        Q(title__icontains=q) | Q(subject__name__icontains=q)
    )

    if scoped and class_id:
        questions = questions.filter(class_obj_id=class_id)
        notes = notes.filter(class_obj_id=class_id)
        papers = papers.filter(class_obj_id=class_id)
    if scoped and board_id:
        questions = questions.filter(board_id=board_id)
        papers = papers.filter(Q(board_id=board_id) | Q(board__isnull=True))

    return {
        'questions': list(questions[:limit]),
        'notes': list(notes[:limit]),
        'papers': list(papers[:limit]),
    }


def search_api(request):
    """JSON endpoint for the ⌘K quick-search modal."""
    query = request.GET.get('q', '')
    scoped = request.GET.get('all') != '1'
    board_id, class_id = _search_scope(request)
    results = _run_search(query, board_id, class_id, scoped=scoped, limit=5)

    def label(obj_subject, obj_class):
        bits = []
        if obj_subject:
            bits.append(obj_subject)
        if obj_class:
            bits.append(obj_class)
        return ' · '.join(bits)

    payload = {
        'questions': [{
            'id': x.pk,
            'title': (x.question_text[:80] + '…') if len(x.question_text) > 80 else x.question_text,
            'meta': label(x.subject.name if x.subject_id else '', x.class_obj.name if x.class_obj_id else '') + (' · ' + str(x.year) if x.year else ''),
            'url': f'/question-bank/?subject={x.subject_id}' if x.subject_id else '/question-bank/',
        } for x in results['questions']],
        'notes': [{
            'id': x.pk,
            'title': x.title,
            'meta': label(x.subject.name if x.subject_id else '', x.class_obj.name if x.class_obj_id else ''),
            'url': f'/study-notes/{x.pk}/',
        } for x in results['notes']],
        'papers': [{
            'id': x.pk,
            'title': x.title,
            'meta': label(x.subject.name if x.subject_id else '', x.class_obj.name if x.class_obj_id else ''),
            'url': f'/exam-papers/{x.pk}/',
        } for x in results['papers']],
    }
    payload['total'] = len(payload['questions']) + len(payload['notes']) + len(payload['papers'])
    payload['scoped'] = scoped
    return JsonResponse(payload)


def search_page(request):
    """Full-page grouped search results."""
    query = request.GET.get('q', '')
    scoped = request.GET.get('all') != '1'
    board_id, class_id = _search_scope(request)
    results = _run_search(query, board_id, class_id, scoped=scoped, limit=30)
    total = len(results['questions']) + len(results['notes']) + len(results['papers'])
    return render(request, 'core/search.html', {
        'query': query,
        'results': results,
        'total': total,
        'scoped': scoped,
    })


SUBJECT_COLOR_HEX = {
    'blue': '#3b82f6', 'red': '#ef4444', 'green': '#22c55e',
    'purple': '#a855f7', 'orange': '#f97316', 'yellow': '#eab308',
    'pink': '#ec4899', 'teal': '#14b8a6', 'indigo': '#6366f1',
    'cyan': '#06b6d4', 'emerald': '#10b981',
}

_SUBJECT_EMOJI = {
    'bangla': '✍️',
    'বাংলা': '✍️',
    'english': '💬',
    'higher math': '📐',
    'higher mathematics': '📐',
    'উচ্চতর গণিত': '📐',
    'physics': '⚛️',
    'পদার্থবিজ্ঞান': '⚛️',
    'chemistry': '🧪',
    'রসায়ন': '🧪',
    'biology': '🧬',
    'জীববিজ্ঞান': '🧬',
    'ict': '💻',
    'information and communication technology': '💻',
    'math': '🔢',
    'mathematics': '🔢',
    'গণিত': '🔢',
    'history': '🏛️',
    'geography': '🌍',
    'economics': '📊',
    'accounting': '🧾',
    'civics': '⚖️',
    'islam': '🕌',
    'religion': '🙏',
}


def _streak_from_dates(active_dates, today):
    """Walk backwards from today (or yesterday) and count consecutive active days."""
    from datetime import timedelta
    if today in active_dates:
        day = today
    elif (today - timedelta(days=1)) in active_dates:
        day = today - timedelta(days=1)
    else:
        return 0
    count = 0
    while day in active_dates:
        count += 1
        day = day - timedelta(days=1)
    return count


def _longest_streak(active_dates):
    from datetime import timedelta
    if not active_dates:
        return 0
    sorted_dates = sorted(active_dates)
    longest = run = 1
    for i in range(1, len(sorted_dates)):
        if (sorted_dates[i] - sorted_dates[i - 1]).days == 1:
            run += 1
            longest = max(longest, run)
        else:
            run = 1
    return max(longest, run)


@login_required
def dashboard(request):
    try:
        if request.user.profile.role == 'ADMIN' and not request.user.profile.is_superadmin:
            return redirect('teacher_dashboard')
    except:
        pass

    try:
        if not request.user.profile.is_premium:
            messages.error(request, _L(request, 'This feature is for Premium users only.', 'এই feature শুধু Premium users এর জন্য।'))
            return redirect('pricing')
    except:
        return redirect('pricing')

    from .models import UserProgress, TeacherFeedback
    from django.db.models import Count, Q
    from django.db.models.functions import TruncDate
    from datetime import timedelta
    from django.utils import timezone
    from collections import defaultdict, Counter

    user = request.user
    progress = UserProgress.objects.filter(user=user).select_related(
        'question', 'question__subject'
    )

    # ----- Headline stats -----
    total_answered = progress.count()
    total_correct = progress.filter(is_correct=True).count()
    total_wrong = total_answered - total_correct
    accuracy = round((total_correct / total_answered * 100), 1) if total_answered else 0

    # ----- This week vs last week (for trend arrows) -----
    today = timezone.now().date()
    week_start = today - timedelta(days=6)
    last_week_start = today - timedelta(days=13)
    last_week_end = today - timedelta(days=7)

    this_week_qs = progress.filter(answered_at__date__gte=week_start)
    last_week_qs = progress.filter(answered_at__date__range=(last_week_start, last_week_end))
    this_week_count = this_week_qs.count()
    last_week_count = last_week_qs.count()
    this_week_correct = this_week_qs.filter(is_correct=True).count()
    last_week_correct = last_week_qs.filter(is_correct=True).count()

    this_week_acc = round((this_week_correct / this_week_count * 100), 1) if this_week_count else 0
    last_week_acc = round((last_week_correct / last_week_count * 100), 1) if last_week_count else 0
    accuracy_delta = round(this_week_acc - last_week_acc, 1)

    if last_week_count == 0:
        volume_delta_pct = None
    else:
        volume_delta_pct = round(((this_week_count - last_week_count) / last_week_count) * 100, 1)

    # ----- Streak -----
    active_dates = set(
        progress.annotate(d=TruncDate('answered_at'))
                .values_list('d', flat=True).distinct()
    )
    current_streak = _streak_from_dates(active_dates, today)
    longest_streak = _longest_streak(active_dates)

    daily_goal = 10
    today_count = progress.filter(answered_at__date=today).count()
    goal_progress = min(round((today_count / daily_goal) * 100), 100) if daily_goal else 0

    lang = getattr(request, 'LANG', 'bn')
    if current_streak >= 30:
        streak_msg = "🔥 You're nearly unstoppable — keep the fire alive!" if lang == 'en' else "🔥 দারুণ! তুমি প্রায় অপ্রতিরোধ্য — এই আগুন ধরে রাখো!"
    elif current_streak >= 7:
        streak_msg = "A full week of consistency — keep going!" if lang == 'en' else "এক সপ্তাহ ধরে ধারাবাহিকভাবে চালিয়ে যাচ্ছ — চালিয়ে যাও!"
    elif current_streak >= 3:
        streak_msg = "Great streak building! Practice a few more questions today." if lang == 'en' else "ভালো ধারাবাহিকতা তৈরি হচ্ছে। আজ আরও কয়েকটা প্রশ্নের অনুশীলন করো।"
    elif current_streak >= 1:
        streak_msg = "You've started! Log in tomorrow to keep your streak alive." if lang == 'en' else "শুরু হয়েছে! আগামীকালও লগ ইন করে ধারাবাহিকতা ধরে রাখো।"
    else:
        streak_msg = "Start today — answer just one question to begin your streak!" if lang == 'en' else "আজই শুরু করো — একটাই প্রশ্নের উত্তর দাও, ধারাবাহিকতা শুরু হবে।"

    # ----- 30-day heatmap -----
    counts_30 = dict(
        progress.filter(answered_at__date__gte=today - timedelta(days=29))
                .annotate(d=TruncDate('answered_at'))
                .values_list('d')
                .annotate(c=Count('id'))
                .values_list('d', 'c')
    )
    heatmap = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        c = counts_30.get(day, 0)
        if c == 0:
            lvl = 0
        elif c < 5:
            lvl = 1
        elif c < 15:
            lvl = 2
        elif c < 30:
            lvl = 3
        else:
            lvl = 4
        heatmap.append({
            'date': day.strftime('%Y-%m-%d'),
            'label': day.strftime('%d %b'),
            'count': c,
            'level': lvl,
        })

    # ----- 7-day daily activity -----
    daily_data = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        c = progress.filter(answered_at__date=day).count()
        cc = progress.filter(answered_at__date=day, is_correct=True).count()
        daily_data.append({'day': day.strftime('%a'), 'count': c, 'correct': cc})

    # ----- 30-day accuracy trend -----
    by_day = {}
    rows_30 = (
        progress.filter(answered_at__date__gte=today - timedelta(days=29))
                .annotate(d=TruncDate('answered_at'))
                .values('d')
                .annotate(total=Count('id'), correct=Count('id', filter=Q(is_correct=True)))
    )
    for r in rows_30:
        by_day[r['d']] = (r['total'], r['correct'])
    accuracy_trend = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        t, c = by_day.get(day, (0, 0))
        accuracy_trend.append({
            'day': day.strftime('%d %b'),
            'accuracy': round((c / t * 100), 1) if t else None,
        })

    # ----- Subjects (collapsible cards + pie) -----
    subj_agg = list(
        progress.values(
            'question__subject__id',
            'question__subject__name',
            'question__subject__icon',
            'question__subject__color',
        ).annotate(
            total=Count('id'),
            correct=Count('id', filter=Q(is_correct=True)),
            written_correct=Count('id', filter=Q(is_correct=True, question__question_type='WRITTEN')),
            mcq_correct=Count('id', filter=Q(is_correct=True, question__question_type='MCQ')),
        ).order_by('-total')
    )
    diff_rows = list(
        progress.values('question__subject__id', 'question__difficulty')
                .annotate(c=Count('id'))
    )
    diff_map = defaultdict(lambda: {'Easy': 0, 'Medium': 0, 'Hard': 0})
    for r in diff_rows:
        diff_map[r['question__subject__id']][r['question__difficulty']] = r['c']

    week_rows = list(
        this_week_qs.annotate(d=TruncDate('answered_at'))
                    .values('question__subject__id', 'd')
                    .annotate(c=Count('id'))
    )
    week_map = defaultdict(dict)
    for r in week_rows:
        week_map[r['question__subject__id']][r['d']] = r['c']

    subjects_data = []
    for s in subj_agg:
        sid = s['question__subject__id']
        total = s['total']
        correct = s['correct']
        wrong = total - correct
        acc = round((correct / total * 100), 1) if total else 0
        weekly = []
        max_in_week = 0
        for i in range(6, -1, -1):
            day = today - timedelta(days=i)
            c = week_map.get(sid, {}).get(day, 0)
            max_in_week = max(max_in_week, c)
            weekly.append({'day': day.strftime('%a'), 'count': c})
        for w in weekly:
            w['height'] = round((w['count'] / max_in_week) * 100) if max_in_week else 0
        color_name = s['question__subject__color'] or 'blue'
        marks = s['mcq_correct'] * 1 + s['written_correct'] * 10
        subjects_data.append({
            'id': sid,
            'name': s['question__subject__name'],
            'icon': s['question__subject__icon'] or _SUBJECT_EMOJI.get(s['question__subject__name'].lower(), '📚'),
            'color': color_name,
            'hex': SUBJECT_COLOR_HEX.get(color_name, '#3b82f6'),
            'total': total,
            'correct': correct,
            'wrong': wrong,
            'accuracy': acc,
            'marks': marks,
            'difficulty': diff_map[sid],
            'weekly': weekly,
        })

    # ----- Insights -----
    weekday_total = Counter()
    for ts in progress.values_list('answered_at', flat=True):
        weekday_total[ts.strftime('%A')] += 1
    best_day = max(weekday_total, key=weekday_total.get) if weekday_total else None

    weakest = None
    for s in subjects_data:
        if s['total'] >= 5:
            if weakest is None or s['accuracy'] < weakest['accuracy']:
                weakest = s

    # ----- Rank (among students) by total correct answers -----
    rank = User.objects.filter(profile__role='STUDENT').annotate(
        cc=Count('progress', filter=Q(progress__is_correct=True))
    ).filter(cc__gt=total_correct).count() + 1
    total_students = User.objects.filter(profile__role='STUDENT').count() or 1

    # ----- Teacher feedback -----
    feedbacks = list(
        TeacherFeedback.objects.filter(student=user)
                                .select_related('teacher', 'progress__question__subject')
                                .order_by('-created_at')[:3]
    )
    unread_count = TeacherFeedback.objects.filter(student=user, is_read=False).count()

    # ----- Exam results grouped by subject -----
    from .models import ExamAttempt

    graded_attempts = list(
        ExamAttempt.objects.filter(student=user, status='GRADED')
                           .select_related('exam_paper', 'exam_paper__subject')
                           .order_by('-graded_at')
    )

    exam_results_by_subject = []
    if graded_attempts:
        subject_groups = {}
        for a in graded_attempts:
            max_marks = EXAM_TOTAL_MAX
            score = a.total_score or 0
            pct = round(score / max_marks * 100, 1) if max_marks else 0

            subj = a.exam_paper.subject
            sid = subj.id if subj else 0
            if sid not in subject_groups:
                color_name = (subj.color if subj else None) or 'blue'
                subject_groups[sid] = {
                    'id': sid,
                    'name': subj.name if subj else 'Other',
                    'icon': (subj.icon if subj and subj.icon
                             else _SUBJECT_EMOJI.get((subj.name if subj else '').lower(), '📚')),
                    'color': color_name,
                    'hex': SUBJECT_COLOR_HEX.get(color_name, '#3b82f6'),
                    'attempts': [],
                    'pct_sum': 0,
                    'best_pct': 0,
                    'count': 0,
                }
            subject_groups[sid]['attempts'].append({
                'id': a.id,
                'title': a.exam_paper.title,
                'score': score,
                'max': max_marks,
                'pct': pct,
                'grade': a.grade,
                'date': a.graded_at,
            })
            subject_groups[sid]['pct_sum'] += pct
            subject_groups[sid]['count'] += 1
            if pct > subject_groups[sid]['best_pct']:
                subject_groups[sid]['best_pct'] = pct

        for s in subject_groups.values():
            s['avg_pct'] = round(s['pct_sum'] / s['count'], 1) if s['count'] else 0
            s['attempts'].sort(key=lambda x: x['date'])
            for att in s['attempts']:
                att['bar_height'] = max(4, min(100, int(att['pct'])))
            exam_results_by_subject.append(s)
        exam_results_by_subject.sort(key=lambda s: -s['count'])

    return render(request, 'core/dashboard.html', {
        # headline
        'total_answered': total_answered,
        'total_correct': total_correct,
        'total_wrong': total_wrong,
        'accuracy': accuracy,
        # trends
        'this_week_count': this_week_count,
        'last_week_count': last_week_count,
        'volume_delta_pct': volume_delta_pct,
        'this_week_acc': this_week_acc,
        'accuracy_delta': accuracy_delta,
        # streak
        'current_streak': current_streak,
        'longest_streak': longest_streak,
        'today_count': today_count,
        'daily_goal': daily_goal,
        'goal_progress': goal_progress,
        'streak_msg': streak_msg,
        'heatmap': heatmap,
        # charts
        'daily_data': daily_data,
        'accuracy_trend': accuracy_trend,
        # subjects
        'subjects_data': subjects_data,
        # insights
        'best_day': best_day,
        'weakest': weakest,
        # rank
        'rank': rank,
        'total_students': total_students,
        # feedback
        'feedbacks': feedbacks,
        'unread_count': unread_count,
        # exam results
        'exam_results_by_subject': exam_results_by_subject,
    })


@login_required
def progress_history(request):
    try:
        if request.user.profile.role == 'ADMIN' and not request.user.profile.is_superadmin:
            return redirect('teacher_dashboard')
    except:
        pass

    try:
        if not request.user.profile.is_premium:
            messages.error(request, _L(request, 'This feature is for Premium users only.', 'এই feature শুধু Premium users এর জন্য।'))
            return redirect('pricing')
    except:
        return redirect('pricing')

    from .models import UserProgress
    from django.db.models import Count, Q
    from django.db.models.functions import TruncDate
    from datetime import timedelta, datetime as dt
    from django.utils import timezone
    from django.core.paginator import Paginator

    user = request.user
    base = UserProgress.objects.filter(user=user).select_related(
        'question', 'question__subject', 'question__board'
    )

    # Filters
    subject_filter = request.GET.get('subject') or ''
    result_filter = request.GET.get('result') or ''
    difficulty_filter = request.GET.get('difficulty') or ''
    date_from = request.GET.get('from') or ''
    date_to = request.GET.get('to') or ''

    qs = base
    if subject_filter:
        qs = qs.filter(question__subject_id=subject_filter)
    if result_filter == 'correct':
        qs = qs.filter(is_correct=True)
    elif result_filter == 'wrong':
        qs = qs.filter(is_correct=False)
    if difficulty_filter in ('Easy', 'Medium', 'Hard'):
        qs = qs.filter(question__difficulty=difficulty_filter)
    if date_from:
        try:
            qs = qs.filter(answered_at__date__gte=dt.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            qs = qs.filter(answered_at__date__lte=dt.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    qs = qs.order_by('-answered_at')

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    filtered_count = paginator.count

    # Summary stats — over user's full history (not the filter)
    today = timezone.now().date()
    week_start = today - timedelta(days=6)
    week_qs = base.filter(answered_at__date__gte=week_start)
    week_total = week_qs.count()
    week_correct = week_qs.filter(is_correct=True).count()
    week_accuracy = round((week_correct / week_total * 100), 1) if week_total else 0

    active_dates = set(
        base.annotate(d=TruncDate('answered_at')).values_list('d', flat=True).distinct()
    )
    current_streak = _streak_from_dates(active_dates, today)

    total_all = base.count()
    correct_all = base.filter(is_correct=True).count()
    accuracy_all = round((correct_all / total_all * 100), 1) if total_all else 0

    subjects = Subject.objects.filter(is_active=True).order_by('name')

    params = request.GET.copy()
    params.pop('page', None)
    qs_string = params.urlencode()

    has_filters = any([subject_filter, result_filter, difficulty_filter, date_from, date_to])

    return render(request, 'core/progress_history.html', {
        'page_obj': page_obj,
        'filtered_count': filtered_count,
        'subjects': subjects,
        'subject_filter': subject_filter,
        'result_filter': result_filter,
        'difficulty_filter': difficulty_filter,
        'date_from': date_from,
        'date_to': date_to,
        'has_filters': has_filters,
        'qs_string': qs_string,
        # summary
        'current_streak': current_streak,
        'week_total': week_total,
        'week_correct': week_correct,
        'week_accuracy': week_accuracy,
        'total_all': total_all,
        'accuracy_all': accuracy_all,
    })


@premium_required
def practical_lab(request):
    from .models import PracticalVideo
    videos = PracticalVideo.objects.filter(is_active=True)
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()

    subject_filter = request.GET.get('subject')
    class_filter = request.GET.get('class_obj')

    if subject_filter:
        videos = videos.filter(subject__slug=subject_filter)
    if class_filter:
        videos = videos.filter(class_obj__id=class_filter)

    return render(request, 'core/practical_lab.html', {
        'videos': videos,
        'subjects': subjects,
        'classes': classes,
    })


def question_bank(request):
    boards = Board.objects.filter(is_active=True)
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    questions = Question.objects.select_related('board', 'subject', 'class_obj').filter(is_active=True)

    board = request.GET.get('board')
    subject = request.GET.get('subject')
    class_id = request.GET.get('class')
    year = request.GET.get('year')
    qtype = request.GET.get('type')                 # MCQ | WRITTEN
    difficulties = request.GET.getlist('difficulty')  # Easy/Medium/Hard (multi)
    status = request.GET.get('status')              # unsolved | wrong

    # Default scope to the student's profile board/class if not overridden
    profile = None
    if request.user.is_authenticated:
        try:
            profile = request.user.profile
        except UserProfile.DoesNotExist:
            profile = None
    if profile and 'board' not in request.GET and getattr(profile, 'board_id', None):
        board = str(profile.board_id)
    if profile and 'class' not in request.GET and getattr(profile, 'class_obj_id', None):
        class_id = str(profile.class_obj_id)

    if board:
        questions = questions.filter(board_id=board)
    if subject:
        questions = questions.filter(subject_id=subject)
    if class_id:
        questions = questions.filter(class_obj_id=class_id)
    if year:
        questions = questions.filter(year=year)
    if qtype in ('MCQ', 'WRITTEN'):
        questions = questions.filter(question_type=qtype)
    if difficulties:
        valid = [d for d in difficulties if d in ('Easy', 'Medium', 'Hard')]
        if valid:
            questions = questions.filter(difficulty__in=valid)

    is_premium = False
    is_teacher = False
    if profile:
        is_premium = profile.is_premium
        is_teacher = profile.role == 'ADMIN' or profile.is_superadmin

    # Personal status filters (logged-in students only)
    if request.user.is_authenticated and not is_teacher and status in ('unsolved', 'wrong'):
        answered_ids = UserProgress.objects.filter(user=request.user).values_list('question_id', flat=True)
        if status == 'unsolved':
            questions = questions.exclude(pk__in=list(answered_ids))
        elif status == 'wrong':
            wrong_ids = UserProgress.objects.filter(user=request.user, is_correct=False).values_list('question_id', flat=True)
            questions = questions.filter(pk__in=list(wrong_ids))

    # Count of active refine filters (for the UI badge)
    active_filters = sum([
        bool(request.GET.get('board')), bool(request.GET.get('class')),
        bool(subject), bool(year), bool(qtype),
        bool(difficulties), bool(status),
    ])
    total_matched = questions.count()

    if not is_premium and not is_teacher:
        questions = questions[:10]

    # Per-question attempt stats for teachers
    if is_teacher:
        from .models import UserProgress
        from django.db.models import Count, Q as DQ
        questions = list(questions)
        q_ids = [q.pk for q in questions]
        stats_map = {}
        for s in UserProgress.objects.filter(question_id__in=q_ids).values('question_id').annotate(
            attempted=Count('id'),
            correct=Count('id', filter=DQ(is_correct=True))
        ):
            pct = round(s['correct'] / s['attempted'] * 100) if s['attempted'] else 0
            stats_map[s['question_id']] = (s['attempted'], pct)
        for q in questions:
            sv = stats_map.get(q.pk, (0, 0))
            q.stats_attempted = sv[0]
            q.stats_correct_pct = sv[1]

    # Map question_id → WrittenSolveSubmission for the current user
    written_solves = {}
    if request.user.is_authenticated and not is_teacher:
        from .models import WrittenSolveSubmission
        q_ids = [q.pk for q in questions]
        for sub in WrittenSolveSubmission.objects.filter(student=request.user, question_id__in=q_ids):
            written_solves[sub.question_id] = sub

    return render(request, 'core/question_bank.html', {
        'boards': boards,
        'subjects': subjects,
        'classes': classes,
        'questions': questions,
        'years': YEARS,
        'is_premium': is_premium,
        'is_teacher': is_teacher,
        'written_solves': written_solves,
        'active_filters': active_filters,
        'total_matched': total_matched,
        'sel_type': qtype or '',
        'sel_difficulties': difficulties,
        'sel_status': status or '',
    })


@login_required
def track_progress(request):
    if request.method == 'POST':
        try:
            profile = request.user.profile
            if profile.role == 'ADMIN' or profile.is_superadmin:
                return JsonResponse({'status': 'ok'})
        except UserProfile.DoesNotExist:
            pass
        from .models import UserProgress
        data = json.loads(request.body)
        question_id = data.get('question_id')
        is_correct = data.get('is_correct', False)
        question_obj = Question.objects.filter(pk=question_id).first()
        if question_obj:
            UserProgress.objects.get_or_create(
                user=request.user,
                question=question_obj,
                defaults={'is_correct': is_correct}
            )
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'})


# -------- SUPERADMIN DASHBOARD --------

@superadmin_required
def superadmin_dashboard(request):
    from .models import PracticalVideo, StudyNote, Contest

    total_superadmins = UserProfile.objects.filter(is_superadmin=True).count()
    total_users = UserProfile.objects.filter(is_superadmin=False).count()
    total_students = UserProfile.objects.filter(role='STUDENT', is_superadmin=False).count()
    total_teachers = UserProfile.objects.filter(role='ADMIN', is_superadmin=False).count()
    free_users = UserProfile.objects.filter(plan='FREE', is_superadmin=False).count()
    basic_users = UserProfile.objects.filter(plan='BASIC', is_superadmin=False).count()
    premium_users = UserProfile.objects.filter(plan='PREMIUM', is_superadmin=False).count()
    paid_users = basic_users + premium_users
    total_questions = Question.objects.filter(is_active=True).count()
    total_boards = Board.objects.filter(is_active=True).count()
    total_subjects = Subject.objects.filter(is_active=True).count()
    total_videos = PracticalVideo.objects.filter(is_active=True).count()
    total_notes = StudyNote.objects.filter(is_active=True).count()
    total_contests = Contest.objects.filter(is_active=True).count()
    recent_users = UserProfile.objects.filter(is_superadmin=False).select_related('user').order_by('-user__date_joined')[:10]
    pending_teachers_count = UserProfile.objects.filter(role='ADMIN', is_approved=False, is_superadmin=False).count()

    return render(request, 'core/superadmin_dashboard.html', {
        'total_superadmins': total_superadmins,
        'total_users': total_users,
        'total_students': total_students,
        'total_teachers': total_teachers,
        'free_users': free_users,
        'basic_users': basic_users,
        'premium_users': premium_users,
        'paid_users': paid_users,
        'total_questions': total_questions,
        'total_boards': total_boards,
        'total_subjects': total_subjects,
        'total_videos': total_videos,
        'total_notes': total_notes,
        'total_contests': total_contests,
        'recent_users': recent_users,
        'pending_teachers_count': pending_teachers_count,
    })


@login_required
def teacher_pending(request):
    try:
        profile = request.user.profile
        if profile.role != 'ADMIN':
            return redirect('home')
        if profile.is_approved or profile.is_superadmin:
            return redirect('teacher_dashboard')
    except Exception:
        return redirect('home')
    return render(request, 'core/teacher_pending.html', {'profile': profile})


@superadmin_required
def teacher_applications(request):
    from .models import Subject
    pending = UserProfile.objects.filter(
        role='ADMIN', is_approved=False, is_superadmin=False
    ).select_related('user').order_by('user__date_joined')
    approved = UserProfile.objects.filter(
        role='ADMIN', is_approved=True, is_superadmin=False
    ).select_related('user').prefetch_related('subjects').order_by('-user__date_joined')[:30]
    all_subjects = Subject.objects.filter(is_active=True).order_by('name')
    return render(request, 'core/teacher_applications.html', {
        'pending': pending,
        'approved': approved,
        'all_subjects': all_subjects,
    })


@superadmin_required
def assign_teacher_subjects(request, pk):
    if request.method == 'POST':
        profile = get_object_or_404(UserProfile, pk=pk, role='ADMIN')
        subject_ids = request.POST.getlist('subjects')
        profile.subjects.set(subject_ids)
        messages.success(request, f'Subjects updated for {profile.user.username}.')
    return redirect('teacher_applications')


@superadmin_required
def approve_teacher(request, pk):
    from django.shortcuts import get_object_or_404
    from .models import Notification
    profile = get_object_or_404(UserProfile, pk=pk, role='ADMIN')
    profile.is_approved = True
    profile.rejection_reason = ''
    profile.save()
    Notification.objects.create(
        recipient=profile.user,
        notif_type='question',
        title='Teacher Application Approved!',
        title_bn='শিক্ষক আবেদন অনুমোদিত!',
        message='Your teacher application has been approved. You can now access your teacher dashboard.',
        message_bn='আপনার শিক্ষক আবেদন অনুমোদিত হয়েছে। এখন আপনি Teacher Dashboard ব্যবহার করতে পারবেন।',
        link='/teacher/dashboard/',
    )
    messages.success(request, f'{profile.user.username} approved successfully.')
    return redirect('teacher_applications')


@superadmin_required
def reject_teacher(request, pk):
    from django.shortcuts import get_object_or_404
    from .models import Notification
    if request.method == 'POST':
        profile = get_object_or_404(UserProfile, pk=pk, role='ADMIN')
        reason = request.POST.get('reason', 'No reason provided.')
        profile.is_approved = False
        profile.rejection_reason = reason
        profile.save()
        Notification.objects.create(
            recipient=profile.user,
            notif_type='question',
            title='Teacher Application Update',
            title_bn='শিক্ষক আবেদনের আপডেট',
            message=f'Your teacher application was not approved. Reason: {reason}',
            message_bn=f'আপনার শিক্ষক আবেদন অনুমোদিত হয়নি। কারণ: {reason}',
        )
        messages.info(request, f'{profile.user.username} application rejected.')
    return redirect('teacher_applications')


# -------- MANAGE PANEL (ADMIN ONLY) --------

@admin_required
def manage_dashboard(request):
    from .models import PracticalVideo, ExamPaper, ExamAttempt, NoteRequest
    total_questions = Question.objects.filter(is_active=True).count()
    total_videos = PracticalVideo.objects.filter(is_active=True).count()
    total_boards = Board.objects.filter(is_active=True).count()
    total_subjects = Subject.objects.filter(is_active=True).count()
    total_classes = Class.objects.count()
    recent_questions = Question.objects.select_related('board', 'subject').order_by('-created_at')[:5]
    exam_paper_count = ExamPaper.objects.count()
    pending_cq_count = ExamAttempt.objects.filter(status='CQ_PENDING').count()
    pending_note_requests = NoteRequest.objects.filter(status='PENDING').select_related('student', 'subject').order_by('-created_at')[:5]
    pending_note_request_count = NoteRequest.objects.filter(status='PENDING').count()
    return render(request, 'manage/dashboard.html', {
        'total_questions': total_questions,
        'total_videos': total_videos,
        'total_boards': total_boards,
        'total_subjects': total_subjects,
        'total_classes': total_classes,
        'recent_questions': recent_questions,
        'exam_paper_count': exam_paper_count,
        'pending_cq_count': pending_cq_count,
        'pending_note_requests': pending_note_requests,
        'pending_note_request_count': pending_note_request_count,
    })


@admin_required
def manage_questions(request):
    questions = Question.objects.select_related('board', 'subject', 'class_obj').order_by('-created_at')
    return render(request, 'manage/questions.html', {'questions': questions})


@admin_required
def question_add(request):
    boards = Board.objects.filter(is_active=True)
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    if request.method == 'POST':
        board_pk = request.POST.get('board')
        subject_pk = request.POST.get('subject')
        class_pk = request.POST.get('class_obj')
        year_val = request.POST.get('year')
        qtype = request.POST.get('question_type') or 'MCQ'
        difficulty = 'Medium'
        chapter = ''

        def _int_or_default(val, default):
            try: return int(val)
            except (TypeError, ValueError): return default

        try:
            year_int = int(year_val)
        except (TypeError, ValueError):
            year_int = None

        # MCQ branch: handle multiple MCQ rows
        if qtype == 'MCQ':
            mcq_texts = request.POST.getlist('mcq_question_text')
            mcq_opt1 = request.POST.getlist('mcq_option1')
            mcq_opt2 = request.POST.getlist('mcq_option2')
            mcq_opt3 = request.POST.getlist('mcq_option3')
            mcq_opt4 = request.POST.getlist('mcq_option4')
            mcq_correct = request.POST.getlist('mcq_correct_option')

            if not any(t.strip() for t in mcq_texts):
                messages.error(request, _L(request, 'Add at least 1 MCQ.', 'কমপক্ষে ১টি MCQ যোগ করুন।'))
                return render(request, 'manage/question_form.html', {
                    'boards': boards, 'subjects': subjects,
                    'classes': classes, 'years': YEARS, 'action': 'Add',
                })

            # Duplicate check: only one MCQ allowed per (board, subject, class, year)
            if board_pk and year_int and subject_pk and class_pk:
                existing = Question.objects.filter(
                    board_id=board_pk, year=year_int,
                    subject_id=subject_pk, class_obj_id=class_pk,
                    question_type='MCQ',
                ).select_related('board', 'subject').first()
                if existing:
                    messages.error(
                        request,
                        f'এই (Board, Subject, Class, Year) এর জন্য একটি MCQ ইতিমধ্যে যোগ করা আছে। '
                        f'Board: {existing.board.name}, Subject: {existing.subject.name}, Year: {existing.year}. '
                        f'প্রতি Board+Subject+Class+Year এ একটিমাত্র MCQ ও একটিমাত্র CQ যোগ করা যায়।'
                    )
                    return render(request, 'manage/question_form.html', {
                        'boards': boards, 'subjects': subjects,
                        'classes': classes, 'years': YEARS, 'action': 'Add',
                        'post': request.POST,
                    })

            # Single shared file for the whole batch — save once and assign path to every MCQ
            shared_file = request.FILES.get('mcq_question_file')
            shared_file_path = None
            if shared_file:
                from django.core.files.storage import default_storage
                shared_file_path = default_storage.save(
                    f'question_mcq/{shared_file.name}', shared_file
                )

            from django.db import IntegrityError
            added = 0
            last_q = None
            for i, text in enumerate(mcq_texts):
                text = text.strip()
                if not text:
                    continue
                try:
                    correct = int(mcq_correct[i]) if i < len(mcq_correct) else 1
                except (ValueError, IndexError):
                    correct = 1
                try:
                    last_q = Question.objects.create(
                        board_id=board_pk, subject_id=subject_pk,
                        class_obj_id=class_pk, year=year_int,
                        chapter=chapter, question_text=text,
                        question_type='MCQ', difficulty=difficulty,
                        option1=mcq_opt1[i] if i < len(mcq_opt1) else '',
                        option2=mcq_opt2[i] if i < len(mcq_opt2) else '',
                        option3=mcq_opt3[i] if i < len(mcq_opt3) else '',
                        option4=mcq_opt4[i] if i < len(mcq_opt4) else '',
                        correct_option=correct,
                        mcq_question_file=shared_file_path or '',
                    )
                    added += 1
                    # DB rule: only ONE MCQ row per (board, subject, class, year) — stop after first
                    break
                except IntegrityError:
                    messages.error(
                        request,
                        'এই (Board, Subject, Class, Year) এর জন্য একটি MCQ ইতিমধ্যে যোগ করা আছে।'
                    )
                    return render(request, 'manage/question_form.html', {
                        'boards': boards, 'subjects': subjects,
                        'classes': classes, 'years': YEARS, 'action': 'Add',
                        'post': request.POST,
                    })

            if added and last_q:
                _notify_all_students(
                    'question',
                    f'New Question Added — {last_q.subject.name}',
                    f'{last_q.subject.name}, {last_q.class_obj.name} (MCQ)',
                    link='/question-bank/',
                    title_bn=f'নতুন প্রশ্ন যোগ হয়েছে — {last_q.subject.name}',
                    message_bn=f'{last_q.subject.name}, {last_q.class_obj.name} (MCQ)',
                )
                messages.success(request, _L(request, 'MCQ question added.', 'MCQ question যোগ করা হয়েছে।'))
            return redirect('manage_questions')

        # WRITTEN (CQ) branch: single question with stimulus + image + hint + solution
        qtext = (request.POST.get('question_text') or '').strip()

        # Duplicate check: only one CQ allowed per (board, subject, class, year)
        if board_pk and year_int and subject_pk and class_pk:
            existing = Question.objects.filter(
                board_id=board_pk, year=year_int,
                subject_id=subject_pk, class_obj_id=class_pk,
                question_type='WRITTEN',
            ).select_related('board', 'subject').first()
            if existing:
                messages.error(
                    request,
                    f'এই (Board, Subject, Class, Year) এর জন্য একটি CQ ইতিমধ্যে যোগ করা আছে। '
                    f'Board: {existing.board.name}, Subject: {existing.subject.name}, Year: {existing.year}. '
                    f'প্রতি Board+Subject+Class+Year এ একটিমাত্র MCQ ও একটিমাত্র CQ যোগ করা যায়।'
                )
                return render(request, 'manage/question_form.html', {
                    'boards': boards, 'subjects': subjects,
                    'classes': classes, 'years': YEARS, 'action': 'Add',
                    'post': request.POST,
                })

        from django.db import IntegrityError
        try:
            q = Question.objects.create(
                board=get_object_or_404(Board, pk=board_pk),
                subject=get_object_or_404(Subject, pk=subject_pk),
                class_obj=get_object_or_404(Class, pk=class_pk),
                year=year_int,
                chapter=chapter,
                question_text=qtext,
                question_type='WRITTEN',
                difficulty=difficulty,
                answer_hint=request.POST.get('answer_hint', ''),
            )
        except IntegrityError:
            messages.error(
                request,
                'এই (Board, Subject, Class, Year) এর জন্য একটি CQ ইতিমধ্যে যোগ করা আছে।'
            )
            return render(request, 'manage/question_form.html', {
                'boards': boards, 'subjects': subjects,
                'classes': classes, 'years': YEARS, 'action': 'Add',
                'post': request.POST,
            })
        if request.FILES.get('stimulus_image'):
            q.stimulus_image = request.FILES['stimulus_image']
            q.save()
        if request.FILES.get('solution_image'):
            q.solution_image = request.FILES['solution_image']
            q.save()
        _notify_all_students(
            'question',
            f'New Question Added — {q.subject.name}',
            f'{q.subject.name}, {q.class_obj.name} ({q.get_question_type_display()})',
            link='/question-bank/',
            title_bn=f'নতুন প্রশ্ন যোগ হয়েছে — {q.subject.name}',
            message_bn=f'{q.subject.name}, {q.class_obj.name} ({q.get_question_type_display()})',
        )
        messages.success(request, 'Question added successfully!')
        return redirect('manage_questions')
    return render(request, 'manage/question_form.html', {
        'boards': boards, 'subjects': subjects,
        'classes': classes, 'years': YEARS, 'action': 'Add'
    })


@admin_required
def question_add_mcq_bulk(request):
    boards = Board.objects.filter(is_active=True).order_by('name')
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    classes = Class.objects.all().order_by('numeric_value')
    years = list(YEARS)

    if request.method == 'POST':
        board_pk = request.POST.get('board')
        subject_pk = request.POST.get('subject')
        class_pk = request.POST.get('class_obj')
        year_val = request.POST.get('year')
        chapter = ''
        difficulty = 'Medium'

        errors = []
        if not board_pk: errors.append(_L(request, 'Select a board.', 'Board বেছে নিন।'))
        if not subject_pk: errors.append(_L(request, 'Select a subject.', 'Subject বেছে নিন।'))
        if not class_pk: errors.append(_L(request, 'Select a class.', 'Class বেছে নিন।'))
        if not year_val: errors.append(_L(request, 'Select a year.', 'Year বেছে নিন।'))

        mcq_texts = request.POST.getlist('mcq_question_text')
        mcq_opt1 = request.POST.getlist('mcq_option1')
        mcq_opt2 = request.POST.getlist('mcq_option2')
        mcq_opt3 = request.POST.getlist('mcq_option3')
        mcq_opt4 = request.POST.getlist('mcq_option4')
        mcq_correct = request.POST.getlist('mcq_correct_option')
        mcq_marks_list = request.POST.getlist('mcq_marks')

        if not any(t.strip() for t in mcq_texts):
            errors.append(_L(request, 'Add at least 1 MCQ.', 'কমপক্ষে ১টি MCQ যোগ করুন।'))

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'manage/question_add_mcq.html', {
                'boards': boards, 'subjects': subjects,
                'classes': classes, 'years': years, 'post': request.POST,
            })

        try:
            year_int = int(year_val)
        except (TypeError, ValueError):
            year_int = None

        # Duplicate check: only one MCQ allowed per (board, subject, class, year)
        existing = Question.objects.filter(
            board_id=board_pk, year=year_int,
            subject_id=subject_pk, class_obj_id=class_pk,
            question_type='MCQ',
        ).select_related('board', 'subject').first()
        if existing:
            messages.error(
                request,
                f'এই (Board, Subject, Class, Year) এর জন্য একটি MCQ ইতিমধ্যে যোগ করা আছে। '
                f'Board: {existing.board.name}, Subject: {existing.subject.name}, Year: {existing.year}. '
                f'প্রতি Board+Subject+Class+Year এ একটিমাত্র MCQ যোগ করা যায়।'
            )
            return render(request, 'manage/question_add_mcq.html', {
                'boards': boards, 'subjects': subjects,
                'classes': classes, 'years': years, 'post': request.POST,
            })

        # Single shared file for the whole batch — save once and assign path to the MCQ row
        shared_file = request.FILES.get('mcq_question_file')
        shared_file_path = None
        if shared_file:
            from django.core.files.storage import default_storage
            shared_file_path = default_storage.save(
                f'question_mcq/{shared_file.name}', shared_file
            )

        from django.db import IntegrityError
        added = 0
        for i, text in enumerate(mcq_texts):
            text = text.strip()
            if not text:
                continue
            try:
                correct = int(mcq_correct[i]) if i < len(mcq_correct) else 1
            except (ValueError, IndexError):
                correct = 1
            try:
                Question.objects.create(
                    board_id=board_pk,
                    subject_id=subject_pk,
                    class_obj_id=class_pk,
                    year=year_int,
                    chapter=chapter,
                    question_text=text,
                    question_type='MCQ',
                    difficulty=difficulty,
                    option1=mcq_opt1[i] if i < len(mcq_opt1) else '',
                    option2=mcq_opt2[i] if i < len(mcq_opt2) else '',
                    option3=mcq_opt3[i] if i < len(mcq_opt3) else '',
                    option4=mcq_opt4[i] if i < len(mcq_opt4) else '',
                    correct_option=correct,
                    mcq_question_file=shared_file_path or '',
                )
                added += 1
                # Only one MCQ row allowed per (board, subject, class, year) — stop after first
                break
            except IntegrityError:
                messages.error(
                    request,
                    'এই (Board, Subject, Class, Year) এর জন্য একটি MCQ ইতিমধ্যে যোগ করা আছে।'
                )
                return render(request, 'manage/question_add_mcq.html', {
                    'boards': boards, 'subjects': subjects,
                    'classes': classes, 'years': years, 'post': request.POST,
                })

        if added:
            messages.success(request, _L(request, 'MCQ question added.', 'MCQ question যোগ করা হয়েছে।'))
        return redirect('manage_questions')

    return render(request, 'manage/question_add_mcq.html', {
        'boards': boards, 'subjects': subjects,
        'classes': classes, 'years': years,
    })


@login_required
def submit_written_solve(request, question_id):
    if request.method != 'POST':
        return redirect('question_bank')
    from .models import WrittenSolveSubmission
    question = get_object_or_404(Question, pk=question_id, question_type='WRITTEN', is_active=True)
    photos = {
        'photo_ka': request.FILES.get('photo_ka'),
        'photo_kha': request.FILES.get('photo_kha'),
        'photo_ga': request.FILES.get('photo_ga'),
        'photo_gha': request.FILES.get('photo_gha'),
    }
    if not all(photos.values()):
        lang = getattr(request, 'LANG', 'bn')
        messages.error(request, 'All 4 parts (ক, খ, গ, ঘ) must be uploaded.' if lang == 'en' else 'সব ৪টি অংশ (ক, খ, গ, ঘ) আপলোড করো।')
        return redirect('question_bank')
    for photo in photos.values():
        err = _upload_error(photo, kind='image', max_mb=10)
        if err:
            messages.error(request, _L(request, *err))
            return redirect('written_question_practice', question_id=question.pk)
    sub, _ = WrittenSolveSubmission.objects.get_or_create(student=request.user, question=question)
    for field, file in photos.items():
        setattr(sub, field, file)
    sub.save()
    UserProgress.objects.get_or_create(user=request.user, question=question, defaults={'is_correct': True})
    lang = getattr(request, 'LANG', 'bn')
    messages.success(request, '✓ Answer submitted! You earned 10 marks.' if lang == 'en' else '✓ উত্তর জমা হয়েছে! ১০ নম্বর পেয়েছো।')
    return redirect('written_question_practice', question_id=question.pk)


@login_required
def written_question_practice(request, question_id):
    from .models import WrittenSolveSubmission
    question = get_object_or_404(Question, pk=question_id, question_type='WRITTEN', is_active=True)
    is_teacher = _is_exam_staff(request.user)
    submission = None
    student_submissions = None
    if is_teacher:
        student_submissions = (
            WrittenSolveSubmission.objects
            .filter(question=question)
            .select_related('student')
            .order_by('-submitted_at')
        )
    else:
        submission = WrittenSolveSubmission.objects.filter(student=request.user, question=question).first()
    return render(request, 'core/written_question_practice.html', {
        'question': question,
        'submission': submission,
        'is_teacher': is_teacher,
        'student_submissions': student_submissions,
    })


@login_required
def upload_question_solution(request, question_id):
    question = get_object_or_404(Question, pk=question_id, question_type='WRITTEN', is_active=True)
    if not _is_exam_staff(request.user):
        messages.error(request, _L(request, 'Only Teachers can do this.', 'শুধু Teacher এই কাজটি করতে পারবে।'))
        return redirect('written_question_practice', question_id=question.pk)
    if request.method == 'POST' and request.FILES.get('solution_image'):
        question.solution_image = request.FILES['solution_image']
        question.save()
        lang = getattr(request, 'LANG', 'bn')
        messages.success(request, '✓ Solution uploaded successfully.' if lang == 'en' else '✓ Solution upload হয়েছে।')
    return redirect('written_question_practice', question_id=question.pk)


@login_required
def delete_question_solution(request, question_id):
    question = get_object_or_404(Question, pk=question_id, question_type='WRITTEN', is_active=True)
    if not _is_exam_staff(request.user):
        messages.error(request, _L(request, 'Only Teachers can do this.', 'শুধু Teacher এই কাজটি করতে পারবে।'))
        return redirect('written_question_practice', question_id=question.pk)
    if request.method == 'POST' and question.solution_image:
        question.solution_image.delete(save=False)
        question.solution_image = None
        question.save()
        lang = getattr(request, 'LANG', 'bn')
        messages.success(request, '✓ Solution deleted.' if lang == 'en' else '✓ Solution মুছে ফেলা হয়েছে।')
    return redirect('written_question_practice', question_id=question.pk)


@login_required
def delete_student_submission(request, submission_id):
    from .models import WrittenSolveSubmission
    sub = get_object_or_404(WrittenSolveSubmission, pk=submission_id)
    if not _is_exam_staff(request.user):
        messages.error(request, _L(request, 'Only Teachers can do this.', 'শুধু Teacher এই কাজটি করতে পারবে।'))
        return redirect('written_question_practice', question_id=sub.question_id)
    question_id = sub.question_id
    if request.method == 'POST':
        for field in ('photo_ka', 'photo_kha', 'photo_ga', 'photo_gha'):
            photo = getattr(sub, field, None)
            if photo:
                photo.delete(save=False)
        UserProgress.objects.filter(user=sub.student, question_id=question_id).delete()
        student_name = sub.student.get_full_name() or sub.student.username
        sub.delete()
        lang = getattr(request, 'LANG', 'bn')
        messages.success(
            request,
            f'✓ Submission of {student_name} deleted.' if lang == 'en'
            else f'✓ {student_name} এর submission মুছে ফেলা হয়েছে।'
        )
    return redirect('written_question_practice', question_id=question_id)


@admin_required
def question_edit(request, pk):
    question = get_object_or_404(Question, pk=pk)
    boards = Board.objects.filter(is_active=True)
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    if request.method == 'POST':
        question.board = get_object_or_404(Board, pk=request.POST.get('board'))
        question.subject = get_object_or_404(Subject, pk=request.POST.get('subject'))
        question.class_obj = get_object_or_404(Class, pk=request.POST.get('class_obj'))
        question.year = request.POST.get('year')
        question.question_type = request.POST.get('question_type')

        if question.question_type == 'MCQ':
            # Edit form uses the row-based naming (first row of the bulk template)
            mcq_texts = request.POST.getlist('mcq_question_text')
            mcq_opt1 = request.POST.getlist('mcq_option1')
            mcq_opt2 = request.POST.getlist('mcq_option2')
            mcq_opt3 = request.POST.getlist('mcq_option3')
            mcq_opt4 = request.POST.getlist('mcq_option4')
            mcq_correct = request.POST.getlist('mcq_correct_option')
            if mcq_texts:
                question.question_text = (mcq_texts[0] or '').strip()
                question.option1 = mcq_opt1[0] if mcq_opt1 else ''
                question.option2 = mcq_opt2[0] if mcq_opt2 else ''
                question.option3 = mcq_opt3[0] if mcq_opt3 else ''
                question.option4 = mcq_opt4[0] if mcq_opt4 else ''
                try:
                    question.correct_option = int(mcq_correct[0]) if mcq_correct else 1
                except (ValueError, IndexError):
                    question.correct_option = 1
        else:
            question.question_text = (request.POST.get('question_text') or '').strip()
            question.answer_hint = request.POST.get('answer_hint', '')
            if request.FILES.get('stimulus_image'):
                question.stimulus_image = request.FILES['stimulus_image']
            if request.FILES.get('solution_image'):
                question.solution_image = request.FILES['solution_image']
        from django.db import IntegrityError
        try:
            question.save()
        except IntegrityError:
            messages.error(
                request,
                'এই question ইতিমধ্যে যোগ করা আছে — একই Board, Subject, Class, Year এ duplicate allowed না।'
            )
            return render(request, 'manage/question_form.html', {
                'question': question, 'boards': boards,
                'subjects': subjects, 'classes': classes,
                'years': YEARS, 'action': 'Edit'
            })
        messages.success(request, 'Question updated!')
        return redirect('manage_questions')
    return render(request, 'manage/question_form.html', {
        'question': question, 'boards': boards,
        'subjects': subjects, 'classes': classes,
        'years': YEARS, 'action': 'Edit'
    })


@admin_required
def question_delete(request, pk):
    question = get_object_or_404(Question, pk=pk)
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Question deleted!')
    return redirect('manage_questions')


@admin_required
def manage_boards(request):
    boards = Board.objects.all()
    return render(request, 'manage/boards.html', {'boards': boards})


@admin_required
def board_add(request):
    if request.method == 'POST':
        Board.objects.create(
            name=request.POST.get('name'),
            student_count=request.POST.get('student_count', ''),
            is_active=True
        )
        messages.success(request, 'Board added!')
    return redirect('manage_boards')


@admin_required
def board_delete(request, pk):
    if request.method == 'POST':
        get_object_or_404(Board, pk=pk).delete()
        messages.success(request, 'Board deleted!')
    return redirect('manage_boards')


@admin_required
def manage_subjects(request):
    subjects = Subject.objects.all()
    return render(request, 'manage/subjects.html', {'subjects': subjects})


@admin_required
def subject_add(request):
    if request.method == 'POST':
        Subject.objects.create(
            name=request.POST.get('name'),
            icon=request.POST.get('icon', ''),
            color=request.POST.get('color', 'blue'),
            is_active=True
        )
        messages.success(request, 'Subject added!')
    return redirect('manage_subjects')


@admin_required
def subject_delete(request, pk):
    if request.method == 'POST':
        get_object_or_404(Subject, pk=pk).delete()
        messages.success(request, 'Subject deleted!')
    return redirect('manage_subjects')


@admin_required
def manage_classes(request):
    classes = Class.objects.all()
    return render(request, 'manage/classes.html', {'classes': classes})


@admin_required
def class_add(request):
    if request.method == 'POST':
        Class.objects.create(
            name=request.POST.get('name'),
            numeric_value=request.POST.get('numeric_value'),
        )
        messages.success(request, 'Class added!')
    return redirect('manage_classes')


@admin_required
def class_delete(request, pk):
    if request.method == 'POST':
        get_object_or_404(Class, pk=pk).delete()
        messages.success(request, 'Class deleted!')
    return redirect('manage_classes')


@premium_required
def practical_videos(request):
    from .models import PracticalVideo
    videos = PracticalVideo.objects.filter(is_active=True)
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()

    subject_filter = request.GET.get('subject')
    class_filter = request.GET.get('class_obj')

    if subject_filter:
        videos = videos.filter(subject__slug=subject_filter)
    if class_filter:
        videos = videos.filter(class_obj__id=class_filter)

    return render(request, 'core/practical_videos.html', {
        'videos': videos,
        'subjects': subjects,
        'classes': classes,
    })


@admin_required
def video_add(request):
    from .models import PracticalVideo
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    if request.method == 'POST':
        title = request.POST.get('title')
        youtube_url = request.POST.get('youtube_url')
        subject_id = request.POST.get('subject')
        class_id = request.POST.get('class_obj')
        description = request.POST.get('description', '')
        PracticalVideo.objects.create(
            title=title,
            youtube_url=youtube_url,
            subject_id=subject_id,
            class_obj_id=class_id,
            description=description,
            is_active=True
        )
        messages.success(request, 'Video added successfully!')
        return redirect('practical_videos')
    return render(request, 'core/video_add.html', {
        'subjects': subjects,
        'classes': classes,
    })


@admin_required
def video_delete(request, pk):
    from .models import PracticalVideo
    video = get_object_or_404(PracticalVideo, pk=pk)
    video.delete()
    messages.success(request, 'Video deleted!')
    return redirect('practical_videos')


@superadmin_required
def update_user(request, pk):
    profile = get_object_or_404(UserProfile, pk=pk)
    if request.method == 'POST':
        profile.role = request.POST.get('role', profile.role)
        profile.plan = request.POST.get('plan', profile.plan)
        profile.save()
        messages.success(request, f'{profile.user.username} updated!')
    return redirect('superadmin_dashboard')


@superadmin_required
def delete_user(request, pk):
    profile = get_object_or_404(UserProfile, pk=pk)
    if request.method == 'POST':
        user = profile.user
        user.delete()
        messages.success(request, 'User deleted!')
    return redirect('superadmin_dashboard')


@superadmin_required
def cancel_subscription(request, pk):
    profile = get_object_or_404(UserProfile, pk=pk)
    if request.method == 'POST':
        profile.plan = 'FREE'
        profile.save()
        messages.success(request, f'{profile.user.username} subscription cancelled!')
    return redirect('superadmin_dashboard')


@admin_required
def teacher_dashboard(request):
    from .models import UserProgress, TeacherFeedback, Contest, ExamPaper, ExamAttempt
    from datetime import timedelta
    from django.utils import timezone
    from django.db.models import Count, Q

    today = timezone.now().date()
    week_ago = today - timedelta(days=7)

    students = UserProfile.objects.filter(
        role='STUDENT', is_superadmin=False
    ).select_related('user').order_by('-user__date_joined')

    all_progress = UserProgress.objects.filter(
        user__profile__role='STUDENT',
        user__profile__is_superadmin=False
    )

    stat_map = {s['user_id']: s for s in all_progress.values('user_id').annotate(
        total=Count('id'), correct=Count('id', filter=Q(is_correct=True))
    )}
    today_map = {s['user_id']: s['count'] for s in all_progress.filter(
        answered_at__date=today).values('user_id').annotate(count=Count('id'))}
    week_map = {s['user_id']: s for s in all_progress.filter(
        answered_at__date__gte=week_ago).values('user_id').annotate(
        total=Count('id'), correct=Count('id', filter=Q(is_correct=True)))}

    student_data = []
    total_answered_all = 0
    active_today = 0
    accuracy_list = []

    for s in students:
        uid = s.user.id
        st = stat_map.get(uid, {'total': 0, 'correct': 0})
        total, correct = st['total'], st['correct']
        today_count = today_map.get(uid, 0)
        wk = week_map.get(uid, {'total': 0, 'correct': 0})
        week_total, week_correct = wk['total'], wk['correct']
        accuracy = round(correct / total * 100, 1) if total > 0 else 0
        week_accuracy = round(week_correct / week_total * 100, 1) if week_total > 0 else 0
        total_answered_all += total
        if today_count > 0:
            active_today += 1
        if total > 0:
            accuracy_list.append(accuracy)
        student_data.append({
            'profile': s,
            'total': total,
            'correct': correct,
            'wrong': total - correct,
            'accuracy': accuracy,
            'today_count': today_count,
            'week_total': week_total,
            'week_accuracy': week_accuracy,
        })

    avg_accuracy = round(sum(accuracy_list) / len(accuracy_list), 1) if accuracy_list else 0

    at_risk = sorted(
        [s for s in student_data if s['week_total'] >= 3 and s['week_accuracy'] < 50],
        key=lambda x: x['week_accuracy']
    )[:5]

    top_performers = sorted(
        [s for s in student_data if s['total'] >= 5],
        key=lambda x: (-x['accuracy'], -x['total'])
    )[:5]

    subject_perf = all_progress.values('question__subject__name').annotate(
        total=Count('id'), correct=Count('id', filter=Q(is_correct=True))
    ).filter(total__gt=0).order_by('-total')
    subject_performance = [
        {'name': sp['question__subject__name'], 'total': sp['total'],
         'correct': sp['correct'],
         'accuracy': round(sp['correct'] / sp['total'] * 100, 1) if sp['total'] > 0 else 0}
        for sp in subject_perf
    ]

    recent_feedbacks = TeacherFeedback.objects.filter(
        teacher=request.user
    ).select_related('student', 'progress__question').order_by('-created_at')[:8]
    feedbacks_this_week = TeacherFeedback.objects.filter(
        teacher=request.user, created_at__date__gte=week_ago).count()
    questions_set = Contest.objects.filter(created_by=request.user).count()

    daily_data = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        daily_data.append({
            'day': day.strftime('%a'),
            'count': all_progress.filter(answered_at__date=day).count(),
            'correct': all_progress.filter(answered_at__date=day, is_correct=True).count(),
        })

    heatmap_days = []
    max_count = 1
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        cnt = all_progress.filter(answered_at__date=day).count()
        heatmap_days.append({'date': day.strftime('%d %b'), 'count': cnt})
        if cnt > max_count:
            max_count = cnt

    insights = []
    inactive_count = sum(1 for s in student_data if s['week_total'] == 0)
    if inactive_count:
        insights.append(_L(request, f"⚠️ {inactive_count} students did not attempt a single question last week.", f"⚠️ {inactive_count} জন student গত সপ্তাহে একটিও প্রশ্ন করেননি।"))
    if subject_performance:
        weakest = min(subject_performance, key=lambda x: x['accuracy'])
        if weakest['accuracy'] < 60:
            insights.append(_L(request, f"📚 Class average accuracy in {weakest['name']} is only {weakest['accuracy']}% — revision needed.", f"📚 {weakest['name']}-এ class-এর গড় accuracy মাত্র {weakest['accuracy']}% — revision দরকার।"))
    if at_risk:
        insights.append(_L(request, f"🔴 {len(at_risk)} students are below 50% this week — talk to them.", f"🔴 {len(at_risk)} জন student এই সপ্তাহে ৫০%-এর নিচে — তাদের সাথে কথা বলুন।"))
    if len(students) > 0 and active_today < len(students) * 0.3:
        insights.append(_L(request, f"📉 Only {active_today} active today — engagement needs a boost.", f"📉 আজ মাত্র {active_today} জন active — engagement বাড়ানো দরকার।"))
    if not insights:
        insights.append(_L(request, "✅ All good. The class is performing well!", "✅ সব ঠিকঠাক আছে। Class ভালো perform করছে!"))

    # ---- Exam Mode data ----
    exam_papers = ExamPaper.objects.filter(is_active=True).select_related(
        'subject', 'class_obj'
    ).order_by('-created_at')

    exam_paper_data = []
    for paper in exam_papers:
        attempts = paper.attempts.all()
        exam_paper_data.append({
            'paper': paper,
            'total': attempts.count(),
            'pending': attempts.filter(status='CQ_PENDING').count(),
            'graded': attempts.filter(status='GRADED').count(),
            'mcq_count': paper.mcqs.count(),
            'cq_count': paper.cqs.count(),
        })

    pending_cq_count = ExamAttempt.objects.filter(status='CQ_PENDING').count()

    urgent_cutoff = timezone.now() - timedelta(hours=24)
    urgent_cq_count = ExamAttempt.objects.filter(
        status='CQ_PENDING', cq_submitted_at__lt=urgent_cutoff
    ).count()

    recent_exam_pending = ExamAttempt.objects.filter(
        status='CQ_PENDING'
    ).select_related('student', 'exam_paper').order_by('cq_submitted_at')[:6]

    recent_exam_graded = ExamAttempt.objects.filter(
        status='GRADED'
    ).select_related('student', 'exam_paper').order_by('-graded_at')[:5]

    # Latest exam attempt per student (for student table column)
    all_student_ids = [s['profile'].user.id for s in student_data]
    latest_attempt_map = {}
    for attempt in ExamAttempt.objects.filter(
        student_id__in=all_student_ids
    ).select_related('exam_paper').order_by('-started_at'):
        if attempt.student_id not in latest_attempt_map:
            latest_attempt_map[attempt.student_id] = attempt

    for s in student_data:
        s['exam_attempt'] = latest_attempt_map.get(s['profile'].user.id)

    if pending_cq_count:
        insights.append(_L(request, f"📝 {pending_cq_count} CQ submissions are still ungraded.", f"📝 {pending_cq_count}টি CQ submission এখনো grade করা হয়নি।"))

    return render(request, 'teacher/dashboard.html', {
        'student_data': student_data,
        'total_answered_all': total_answered_all,
        'active_today': active_today,
        'avg_accuracy': avg_accuracy,
        'daily_data': daily_data,
        'at_risk': at_risk,
        'at_risk_count': len(at_risk),
        'top_performers': top_performers,
        'subject_performance': subject_performance,
        'recent_feedbacks': recent_feedbacks,
        'feedbacks_this_week': feedbacks_this_week,
        'questions_set': questions_set,
        'heatmap_days': heatmap_days,
        'max_heatmap': max_count,
        'insights': insights,
        'exam_paper_data': exam_paper_data,
        'pending_cq_count': pending_cq_count,
        'urgent_cq_count': urgent_cq_count,
        'recent_exam_pending': recent_exam_pending,
        'recent_exam_graded': recent_exam_graded,
    })


@admin_required
def student_detail(request, pk):
    from .models import UserProgress, TeacherFeedback
    from django.db.models import Count, Q
    from datetime import timedelta
    from django.utils import timezone

    profile = get_object_or_404(UserProfile, pk=pk)
    progress = UserProgress.objects.filter(user=profile.user).select_related('question', 'question__subject')

    total_answered = progress.count()
    total_correct = progress.filter(is_correct=True).count()
    total_wrong = total_answered - total_correct
    accuracy = round(total_correct / total_answered * 100, 1) if total_answered > 0 else 0

    subject_progress = list(progress.values('question__subject__name').annotate(
        total=Count('id'), correct=Count('id', filter=Q(is_correct=True))
    ).order_by('-total'))
    for sp in subject_progress:
        sp['accuracy'] = round(sp['correct'] / sp['total'] * 100, 1) if sp['total'] > 0 else 0

    difficulty_data = {}
    for diff in ['Easy', 'Medium', 'Hard']:
        dq = progress.filter(question__difficulty=diff)
        d_total = dq.count()
        d_correct = dq.filter(is_correct=True).count()
        difficulty_data[diff] = {
            'total': d_total,
            'correct': d_correct,
            'accuracy': round(d_correct / d_total * 100, 1) if d_total > 0 else 0,
        }

    today = timezone.now().date()
    week_ago = today - timedelta(days=7)

    daily_data = []
    for i in range(13, -1, -1):
        day = today - timedelta(days=i)
        daily_data.append({
            'day': day.strftime('%d %b'),
            'count': progress.filter(answered_at__date=day).count(),
            'correct': progress.filter(answered_at__date=day, is_correct=True).count(),
        })

    heatmap_days = []
    max_count = 1
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        cnt = progress.filter(answered_at__date=day).count()
        heatmap_days.append({'date': day.strftime('%d %b'), 'count': cnt})
        if cnt > max_count:
            max_count = cnt

    streak = 0
    check_day = today
    while progress.filter(answered_at__date=check_day).exists():
        streak += 1
        check_day -= timedelta(days=1)

    all_stats = list(UserProgress.objects.filter(
        user__profile__role='STUDENT', user__profile__is_superadmin=False
    ).values('user_id').annotate(
        total=Count('id'), correct=Count('id', filter=Q(is_correct=True))
    ))
    ranked = sorted(
        [(s['user_id'], round(s['correct'] / s['total'] * 100, 1) if s['total'] > 0 else 0)
         for s in all_stats], key=lambda x: -x[1]
    )
    class_rank = next((i + 1 for i, (uid, _) in enumerate(ranked) if uid == profile.user.id), None)
    total_students = UserProfile.objects.filter(role='STUDENT', is_superadmin=False).count()

    week_progress = progress.filter(answered_at__date__gte=week_ago)
    week_total = week_progress.count()
    week_correct = week_progress.filter(is_correct=True).count()
    week_accuracy = round(week_correct / week_total * 100, 1) if week_total > 0 else 0
    is_at_risk = week_total >= 3 and week_accuracy < 50

    insights = []
    if subject_progress:
        weakest = min(subject_progress, key=lambda x: x['accuracy'])
        if weakest['total'] >= 3 and weakest['accuracy'] < 60:
            insights.append(_L(request, f"⚠️ Only {weakest['accuracy']}% accuracy in {weakest['question__subject__name']} — this subject needs special attention.", f"⚠️ {weakest['question__subject__name']}-এ মাত্র {weakest['accuracy']}% accuracy — এই বিষয়ে বিশেষ মনোযোগ দরকার।"))
    easy_acc = difficulty_data['Easy']['accuracy']
    hard_acc = difficulty_data['Hard']['accuracy']
    if difficulty_data['Hard']['total'] >= 3 and difficulty_data['Easy']['total'] >= 3:
        if easy_acc >= 70 and hard_acc < 50:
            insights.append(_L(request, f"Strong on easy ({easy_acc}%) but weak on hard ({hard_acc}%) — move from Medium toward Hard.", f"সহজ প্রশ্নে ভালো ({easy_acc}%) কিন্তু কঠিনে দুর্বল ({hard_acc}%) — Medium থেকে Hard-এ নিয়ে যান।"))
    recent_7 = sum(d['count'] for d in daily_data[-7:])
    prev_7 = sum(d['count'] for d in daily_data[:7])
    if prev_7 > 0 and recent_7 < prev_7 * 0.5:
        insights.append(_L(request, "📉 Engagement dropped notably this week — send motivational feedback.", "📉 এই সপ্তাহে engagement উল্লেখযোগ্যভাবে কমেছে — motivational feedback পাঠান।"))
    elif recent_7 > prev_7 * 1.5 and recent_7 > 0:
        insights.append(_L(request, "📈 Very active this week! Do not forget to encourage them.", "📈 এই সপ্তাহে দারুণ active! উৎসাহ দিতে ভুলবেন না।"))
    if streak == 0:
        insights.append(_L(request, "❌ No questions attempted today. Consider sending a reminder.", "❌ আজ কোনো প্রশ্ন করেনি। Reminder পাঠানো যেতে পারে।"))
    elif streak >= 7:
        insights.append(_L(request, f"🔥 Great {streak}-day streak! Acknowledge it.", f"🔥 {streak} দিনের দারুণ streak! এটা acknowledge করুন।"))
    if not insights:
        insights.append(_L(request, "📊 All good. Keep giving regular feedback.", "📊 সব ঠিকঠাক আছে। Regular feedback দিতে থাকুন।"))

    teacher_feedbacks = TeacherFeedback.objects.filter(
        teacher=request.user, student=profile.user
    ).select_related('progress__question').order_by('-created_at')[:20]

    history = progress.prefetch_related('feedbacks__teacher').order_by('-answered_at')[:30]

    return render(request, 'teacher/student_detail.html', {
        'profile': profile,
        'total_answered': total_answered,
        'total_correct': total_correct,
        'total_wrong': total_wrong,
        'accuracy': accuracy,
        'subject_progress': subject_progress,
        'difficulty_data': difficulty_data,
        'daily_data': daily_data,
        'heatmap_days': heatmap_days,
        'max_heatmap': max_count,
        'streak': streak,
        'class_rank': class_rank,
        'total_students': total_students,
        'is_at_risk': is_at_risk,
        'week_accuracy': week_accuracy,
        'week_total': week_total,
        'insights': insights,
        'teacher_feedbacks': teacher_feedbacks,
        'history': history,
    })


@admin_required
def give_feedback(request, progress_pk):
    from .models import UserProgress, TeacherFeedback
    progress = get_object_or_404(UserProgress, pk=progress_pk)
    if request.method == 'POST':
        comment = request.POST.get('comment', '').strip()
        if comment:
            TeacherFeedback.objects.create(
                teacher=request.user,
                student=progress.user,
                progress=progress,
                comment=comment
            )
            messages.success(request, _L(request, 'Feedback sent!', 'Feedback পাঠানো হয়েছে!'))
    return redirect('student_detail', pk=progress.user.profile.pk)


@admin_required
def send_general_feedback(request, student_pk):
    from .models import UserProgress, TeacherFeedback
    profile = get_object_or_404(UserProfile, pk=student_pk)
    if request.method == 'POST':
        comment = request.POST.get('comment', '').strip()
        if comment:
            latest = UserProgress.objects.filter(user=profile.user).order_by('-answered_at').first()
            if latest:
                TeacherFeedback.objects.create(
                    teacher=request.user,
                    student=profile.user,
                    progress=latest,
                    comment=comment
                )
                messages.success(request, _L(request, 'Feedback sent!', 'Feedback পাঠানো হয়েছে!'))
            else:
                messages.warning(request, _L(request, 'The student has not attempted any questions yet.', 'Student এখনো কোনো প্রশ্ন করেননি।'))
    return redirect('student_detail', pk=student_pk)


@login_required
def request_note(request):
    if request.method != 'POST':
        return redirect('study_notes')
    from .models import NoteRequest, Notification, UserProfile
    topic = request.POST.get('topic', '').strip()
    if not topic:
        messages.error(request, 'Topic দাও।' if request.LANG == 'bn' else 'Please enter a topic.')
        return redirect('study_notes')
    subject_id = request.POST.get('subject')
    subject = Subject.objects.filter(pk=subject_id).first() if subject_id else None
    note_req = NoteRequest.objects.create(
        student=request.user,
        subject=subject,
        topic=topic,
        details=request.POST.get('details', ''),
    )
    admin_ids = UserProfile.objects.filter(is_admin=True).values_list('user_id', flat=True)
    Notification.objects.bulk_create([
        Notification(
            recipient_id=uid,
            notif_type='request',
            title=f'নতুন Note Request: {topic}' if request.LANG == 'bn' else f'New Note Request: {topic}',
            message=f'{request.user.username} একটি note request করেছেন — "{topic}"' if request.LANG == 'bn' else f'{request.user.username} requested a note on "{topic}"',
            link='/manage/note-requests/',
        )
        for uid in admin_ids
    ])
    messages.success(request, 'Request পাঠানো হয়েছে! 📩' if request.LANG == 'bn' else 'Request sent! 📩')
    return redirect('study_notes')


@admin_required
def manage_note_requests(request):
    from .models import NoteRequest
    note_requests = NoteRequest.objects.select_related('student', 'subject', 'fulfilled_by', 'fulfilled_note').order_by('-created_at')
    subjects = Subject.objects.filter(is_active=True)
    return render(request, 'manage/note_requests.html', {
        'note_requests': note_requests,
        'subjects': subjects,
        'pending_count': note_requests.filter(status='PENDING').count(),
    })


@admin_required
def fulfill_note_request(request, pk):
    from .models import NoteRequest, Notification
    from django.utils import timezone
    note_req = get_object_or_404(NoteRequest, pk=pk)
    action = request.POST.get('action', 'fulfill')
    if action == 'reject':
        note_req.status = 'REJECTED'
        note_req.fulfilled_at = timezone.now()
        note_req.fulfilled_by = request.user
        note_req.save()
        Notification.objects.create(
            recipient=note_req.student,
            notif_type='request',
            title='Note Request Update',
            message=f'Your request for "{note_req.topic}" could not be fulfilled at this time.',
            title_bn='Note Request আপডেট',
            message_bn=f'"{note_req.topic}" বিষয়ে তোমার request টি এই মুহূর্তে পূরণ করা সম্ভব হয়নি।',
            link='/student/notifications/',
        )
        messages.info(request, 'Request reject করা হয়েছে।' if request.LANG == 'bn' else 'Request rejected.')
    else:
        note_req.status = 'FULFILLED'
        note_req.fulfilled_at = timezone.now()
        note_req.fulfilled_by = request.user
        note_req.save()
        Notification.objects.create(
            recipient=note_req.student,
            notif_type='note',
            title='Your Note Request is Fulfilled! 🎉',
            message=f'A study note on "{note_req.topic}" has been added. Check it out!',
            title_bn='তোমার Note Request পূরণ হয়েছে! 🎉',
            message_bn=f'"{note_req.topic}" বিষয়ে study note যোগ করা হয়েছে। এখনই দেখো!',
            link='/study-notes/',
        )
        messages.success(request, 'Request fulfilled করা হয়েছে।' if request.LANG == 'bn' else 'Request fulfilled.')
    return redirect('manage_note_requests')


@login_required
def notifications(request):
    from .models import TeacherFeedback, Notification
    feedbacks = TeacherFeedback.objects.filter(
        student=request.user
    ).select_related('teacher', 'progress__question')
    feedbacks.filter(is_read=False).update(is_read=True)

    notifs = Notification.objects.filter(recipient=request.user)
    notifs.filter(is_read=False).update(is_read=True)

    return render(request, 'core/notifications.html', {
        'feedbacks': feedbacks,
        'notifications': notifs,
    })


# -------- STUDY NOTES --------

@premium_required
def study_notes(request):
    from .models import StudyNote, NoteBookmark
    notes = StudyNote.objects.filter(is_active=True).select_related('subject', 'class_obj', 'created_by')
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()

    subject_filter = request.GET.get('subject')
    class_filter = request.GET.get('class_obj')
    search = request.GET.get('search', '')

    if subject_filter:
        notes = notes.filter(subject__slug=subject_filter)
    if class_filter:
        notes = notes.filter(class_obj__id=class_filter)
    if search:
        notes = notes.filter(title__icontains=search) | notes.filter(chapter__icontains=search)

    bookmarked_ids = set()
    if request.user.is_authenticated:
        bookmarked_ids = set(NoteBookmark.objects.filter(
            user=request.user
        ).values_list('note_id', flat=True))

    return render(request, 'core/study_notes.html', {
        'notes': notes,
        'subjects': subjects,
        'classes': classes,
        'search': search,
        'bookmarked_ids': bookmarked_ids,
    })


@premium_required
def study_note_detail(request, pk):
    from .models import StudyNote, NoteBookmark, NoteReadProgress, NoteComment

    note = get_object_or_404(StudyNote, pk=pk, is_active=True)
    is_bookmarked = NoteBookmark.objects.filter(user=request.user, note=note).exists()
    read_progress, _ = NoteReadProgress.objects.get_or_create(user=request.user, note=note)
    related_notes = StudyNote.objects.filter(
        subject=note.subject, is_active=True
    ).exclude(pk=note.pk)[:3]

    approved_comments = NoteComment.objects.filter(note=note, is_approved=True).select_related('user')

    try:
        is_teacher = request.user.profile.role == 'ADMIN'
    except:
        is_teacher = False

    pending_comments = NoteComment.objects.filter(
        note=note, is_approved=False
    ).select_related('user') if is_teacher else NoteComment.objects.none()

    return render(request, 'core/study_note_detail.html', {
        'note': note,
        'is_bookmarked': is_bookmarked,
        'read_progress': read_progress,
        'related_notes': related_notes,
        'approved_comments': approved_comments,
        'pending_comments': pending_comments,
        'is_teacher': is_teacher,
    })


@admin_required
def study_note_add(request):
    from .models import StudyNote, NoteRequest
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    linked_request = None
    req_id = request.GET.get('req') or request.POST.get('note_request_id')
    if req_id:
        linked_request = NoteRequest.objects.filter(pk=req_id, status='PENDING').first()
    if request.method == 'POST':
        err = _upload_error(request.FILES.get('pdf_file'), kind='doc', max_mb=20)
        if err:
            messages.error(request, _L(request, *err))
            return redirect('study_note_add')
        note = StudyNote.objects.create(
            title=request.POST.get('title'),
            subject=get_object_or_404(Subject, pk=request.POST.get('subject')),
            class_obj=get_object_or_404(Class, pk=request.POST.get('class_obj')),
            chapter=request.POST.get('chapter'),
            content=request.POST.get('content', ''),
            created_by=request.user,
            is_active=True
        )
        if request.FILES.get('pdf_file'):
            note.pdf_file = request.FILES['pdf_file']
            note.save()
        _notify_all_students(
            'note',
            f'New Study Note: {note.title}',
            f'{request.user.username} uploaded a new study note — "{note.title}" ({note.subject.name}, {note.class_obj.name})',
            link=f'/study-notes/{note.pk}/',
            title_bn=f'নতুন Study Note: {note.title}',
            message_bn=f'{request.user.username} একটি নতুন study note আপলোড করেছেন — "{note.title}" ({note.subject.name}, {note.class_obj.name})',
        )
        # Fulfill a linked note request if specified
        from .models import NoteRequest
        from django.utils import timezone
        req_id = request.POST.get('note_request_id')
        if req_id:
            try:
                note_req = NoteRequest.objects.get(pk=req_id, status='PENDING')
                note_req.status = 'FULFILLED'
                note_req.fulfilled_at = timezone.now()
                note_req.fulfilled_by = request.user
                note_req.fulfilled_note = note
                note_req.save()
                from .models import Notification
                Notification.objects.create(
                    recipient=note_req.student,
                    notif_type='note',
                    title='Your Note Request is Fulfilled! 🎉',
                    message=f'A new study note on "{note_req.topic}" has been added.',
                    title_bn='তোমার Note Request পূরণ হয়েছে! 🎉',
                    message_bn=f'"{note_req.topic}" বিষয়ে একটি নতুন study note যোগ করা হয়েছে।',
                    link=f'/study-notes/{note.pk}/',
                )
            except NoteRequest.DoesNotExist:
                pass
        messages.success(request, 'Note added!')
        return redirect('study_notes')
    return render(request, 'core/study_note_add.html', {
        'subjects': subjects,
        'classes': classes,
        'linked_request': linked_request,
    })


@admin_required
def study_note_edit(request, pk):
    from .models import StudyNote
    note = get_object_or_404(StudyNote, pk=pk)
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    if request.method == 'POST':
        note.title = request.POST.get('title')
        note.subject = get_object_or_404(Subject, pk=request.POST.get('subject'))
        note.class_obj = get_object_or_404(Class, pk=request.POST.get('class_obj'))
        note.chapter = request.POST.get('chapter')
        note.content = request.POST.get('content', '')
        if request.FILES.get('pdf_file'):
            err = _upload_error(request.FILES['pdf_file'], kind='doc', max_mb=20)
            if err:
                messages.error(request, _L(request, *err))
                return redirect('study_note_edit', pk=note.pk)
            note.pdf_file = request.FILES['pdf_file']
        note.save()
        messages.success(request, 'Note updated!')
        return redirect('study_note_detail', pk=note.pk)
    return render(request, 'core/study_note_edit.html', {
        'note': note,
        'subjects': subjects,
        'classes': classes,
    })


@admin_required
def study_note_delete(request, pk):
    from .models import StudyNote
    note = get_object_or_404(StudyNote, pk=pk)
    if request.method == 'POST':
        note.delete()
        messages.success(request, 'Note deleted!')
    return redirect('study_notes')


@login_required
def toggle_bookmark(request, pk):
    from .models import StudyNote, NoteBookmark
    note = get_object_or_404(StudyNote, pk=pk)
    bookmark, created = NoteBookmark.objects.get_or_create(user=request.user, note=note)
    if not created:
        bookmark.delete()
        return JsonResponse({'bookmarked': False})
    return JsonResponse({'bookmarked': True})


@login_required
def update_read_progress(request, pk):
    from .models import StudyNote, NoteReadProgress
    if request.method == 'POST':
        note = get_object_or_404(StudyNote, pk=pk)
        data = json.loads(request.body)
        scroll_percent = data.get('scroll_percent', 0)
        progress, _ = NoteReadProgress.objects.get_or_create(user=request.user, note=note)
        progress.scroll_percent = scroll_percent
        if scroll_percent >= 90:
            progress.is_completed = True
        progress.save()
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'})


@login_required
def add_comment(request, pk):
    from .models import StudyNote, NoteComment
    note = get_object_or_404(StudyNote, pk=pk)
    if request.method == 'POST':
        comment_text = request.POST.get('comment', '').strip()
        if comment_text:
            is_approved = request.user.profile.role == 'ADMIN' or request.user.profile.is_superadmin
            NoteComment.objects.create(
                note=note,
                user=request.user,
                comment=comment_text,
                is_approved=is_approved
            )
            messages.success(request, 'Comment posted!' if is_approved else 'Comment submitted for approval!')
    return redirect('study_note_detail', pk=pk)


@admin_required
def approve_comment(request, comment_pk):
    from .models import NoteComment
    comment = get_object_or_404(NoteComment, pk=comment_pk)
    if request.method == 'POST':
        comment.is_approved = True
        comment.save()
    return redirect('study_note_detail', pk=comment.note.pk)


@admin_required
def delete_comment(request, comment_pk):
    from .models import NoteComment
    comment = get_object_or_404(NoteComment, pk=comment_pk)
    note_pk = comment.note.pk
    if request.method == 'POST':
        comment.delete()
    return redirect('study_note_detail', pk=note_pk)


@admin_required
def generate_note_ai(request):
    if request.method == 'POST':
        topic = request.POST.get('topic', '')
        subject_id = request.POST.get('subject')
        class_id = request.POST.get('class_obj')
        chapter = request.POST.get('chapter', '')

        prompt = f"এই topic এর উপর একটি সম্পূর্ণ study note বাংলায় লিখো। Note টি SSC/HSC students এর জন্য। Topic: {topic}, Chapter: {chapter}. Note এ heading, subheading, examples, এবং key points থাকবে। HTML format এ দাও। Math/chemistry equation এর জন্য LaTeX delimiter use করো: inline math এর জন্য $...$ এবং display equation এর জন্য $$...$$ (যেমন: $E=mc^2$ অথবা $$\\int_0^1 x\\,dx$$)।"

        try:
            content = ai_svc.anthropic_complete(prompt, max_tokens=2000)
            from .models import StudyNote
            note = StudyNote.objects.create(
                title=topic,
                subject=get_object_or_404(Subject, pk=subject_id),
                class_obj=get_object_or_404(Class, pk=class_id),
                chapter=chapter,
                content=content,
                created_by=request.user,
                is_active=True
            )
            messages.success(request, _L(request, 'Note generated with AI!', 'AI দিয়ে note তৈরি হয়েছে!'))
            return redirect('study_note_detail', pk=note.pk)
        except ai_svc.AIServiceError as e:
            messages.error(request, f'AI error: {e}')

    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    return render(request, 'core/generate_note.html', {
        'subjects': subjects,
        'classes': classes,
    })


@login_required
def generate_mcq(request, pk):
    if request.method == 'POST':
        from .models import StudyNote

        note = get_object_or_404(StudyNote, pk=pk)

        prompt = f"এই study note থেকে ১০টি MCQ প্রশ্ন বাংলায় তৈরি করো। প্রতিটি প্রশ্নে ৪টি option এবং সঠিক উত্তর দাও। Math/chemistry equation এর জন্য LaTeX delimiter use করো ($...$ inline এর জন্য, $$...$$ display এর জন্য)। JSON format এ দাও এভাবে: {{\"mcqs\": [{{\"question\": \"...\", \"options\": [\"ক) ...\", \"খ) ...\", \"গ) ...\", \"ঘ) ...\"], \"answer\": \"ক\"}}]}}\n\nNote:\n{note.content[:3000]}"

        try:
            text = ai_svc.anthropic_complete(prompt, max_tokens=2000)
            import re
            json_match = re.search(r'\{.*\}', text, re.DOTALL)
            if json_match:
                mcq_data = json.loads(json_match.group())
                return JsonResponse({'mcqs': mcq_data.get('mcqs', [])})
            return JsonResponse({'mcqs': [], 'error': 'Could not parse MCQs'})
        except ai_svc.AIServiceError as e:
            return JsonResponse({'error': str(e)})

    return JsonResponse({'error': 'Invalid request'})


@login_required
def summarize_note(request, pk):
    if request.method == 'POST':
        from .models import StudyNote

        note = get_object_or_404(StudyNote, pk=pk)

        prompt = f"এই study note টি সহজ বাংলায় সংক্ষেপ করো। Key points bullet points এ দাও। ৩-৫ টি main point এবং একটি summary paragraph লিখো। Math/chemistry equation এর জন্য LaTeX delimiter use করো ($...$ inline, $$...$$ display)।\n\nNote:\n{note.content[:3000]}"

        try:
            summary = ai_svc.anthropic_complete(prompt, max_tokens=1000)
            return JsonResponse({'summary': summary})
        except ai_svc.AIServiceError as e:
            return JsonResponse({'error': str(e)})

    return JsonResponse({'error': 'Invalid request'})


@login_required
def ask_ai(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        question = data.get('question', '')
        note_content = data.get('note_content', '')

        prompt = f"Based on this study note, answer in Bengali:\n\nNote:\n{note_content}\n\nQuestion: {question}\n\nPlease respond in Bengali. For any math or chemistry equations, wrap them in LaTeX delimiters: $...$ for inline, $$...$$ for display equations."

        try:
            answer = ai_svc.anthropic_complete(prompt, max_tokens=1000)
            return JsonResponse({'answer': answer})
        except ai_svc.AIServiceError as e:
            return JsonResponse({'answer': f'Error: {e}'})

    return JsonResponse({'error': 'Invalid request'})


# -------- CONTEST --------

@login_required
def contest_list(request):
    from .models import Contest, ContestSubmission, ContestRegistration, UserRating
    from django.utils import timezone
    from django.db.models import Count
    from django.core.paginator import Paginator

    now = timezone.now()
    tab = request.GET.get('tab', 'upcoming')
    q = (request.GET.get('q') or '').strip()
    subject_id = request.GET.get('subject')
    difficulty = request.GET.get('difficulty')
    ctype = request.GET.get('type')

    base = Contest.objects.filter(is_active=True).select_related(
        'subject', 'class_obj', 'created_by'
    )
    if tab == 'live':
        base = base.filter(start_time__lte=now, end_time__gte=now)
    elif tab == 'past':
        base = base.filter(end_time__lt=now).order_by('-end_time')
    elif tab == 'mine':
        reg_ids = ContestRegistration.objects.filter(
            user=request.user
        ).values_list('contest_id', flat=True)
        sub_ids = ContestSubmission.objects.filter(
            student=request.user
        ).values_list('contest_id', flat=True)
        base = base.filter(id__in=list(reg_ids) + list(sub_ids))
    elif tab == 'virtual':
        base = base.filter(end_time__lt=now, allows_virtual=True).order_by('-end_time')
    else:
        tab = 'upcoming'
        base = base.filter(start_time__gte=now).order_by('start_time')

    if tab not in ('past', 'virtual'):
        base = base.order_by('start_time')

    if q:
        base = base.filter(title__icontains=q)
    if subject_id:
        base = base.filter(subject_id=subject_id)
    if difficulty:
        base = base.filter(difficulty=difficulty)
    if ctype:
        base = base.filter(contest_type=ctype)

    paginator = Paginator(base, 20)
    page = paginator.get_page(request.GET.get('page'))

    contest_ids = [c.id for c in page.object_list]
    my_registered = set(ContestRegistration.objects.filter(
        user=request.user, contest_id__in=contest_ids,
    ).values_list('contest_id', flat=True))
    my_submitted = set(ContestSubmission.objects.filter(
        student=request.user, is_submitted=True, contest_id__in=contest_ids,
    ).values_list('contest_id', flat=True))

    participant_counts = {
        c['contest_id']: c['cnt']
        for c in ContestRegistration.objects.filter(
            contest_id__in=contest_ids,
        ).values('contest_id').annotate(cnt=Count('id'))
    }

    featured = Contest.objects.filter(
        is_active=True, is_featured=True, end_time__gte=now,
    ).select_related('subject', 'class_obj').first()

    total_contests = Contest.objects.filter(is_active=True).count()
    live_now = Contest.objects.filter(
        is_active=True, start_time__lte=now, end_time__gte=now,
    ).count()
    total_participants = ContestRegistration.objects.values('user').distinct().count()
    my_rating_obj, _ = UserRating.objects.get_or_create(user=request.user)

    return render(request, 'core/contest_list.html', {
        'contests': page.object_list,
        'page': page,
        'paginator': paginator,
        'tab': tab,
        'q': q,
        'subject_id': int(subject_id) if subject_id and subject_id.isdigit() else None,
        'difficulty': difficulty or '',
        'ctype': ctype or '',
        'subjects': Subject.objects.filter(is_active=True),
        'CONTEST_TYPE': Contest.CONTEST_TYPE,
        'DIFFICULTY': Contest.DIFFICULTY,
        'my_registered': my_registered,
        'my_submissions': my_submitted,
        'participant_counts': participant_counts,
        'featured': featured,
        'stats': {
            'total': total_contests,
            'live': live_now,
            'participants': total_participants,
        },
        'my_rating': my_rating_obj,
        'now': now,
    })


@admin_required
def contest_create(request):
    from .models import Contest, ContestQuestion
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    if request.method == 'POST':
        from django.utils import timezone
        import datetime
        contest = Contest.objects.create(
            title=request.POST.get('title'),
            created_by=request.user,
            subject=get_object_or_404(Subject, pk=request.POST.get('subject')),
            class_obj=get_object_or_404(Class, pk=request.POST.get('class_obj')),
            duration_minutes=int(request.POST.get('duration_minutes', 30)),
            start_time=request.POST.get('start_time'),
            end_time=request.POST.get('end_time'),
            is_active=True
        )
        question_texts = request.POST.getlist('question_text')
        question_types = request.POST.getlist('question_type')
        marks_list = request.POST.getlist('marks')
        option1s = request.POST.getlist('option1')
        option2s = request.POST.getlist('option2')
        option3s = request.POST.getlist('option3')
        option4s = request.POST.getlist('option4')
        correct_options = request.POST.getlist('correct_option')

        for i, qtext in enumerate(question_texts):
            if qtext.strip():
                ContestQuestion.objects.create(
                    contest=contest,
                    question_text=qtext,
                    question_type=question_types[i] if i < len(question_types) else 'MCQ',
                    marks=int(marks_list[i]) if i < len(marks_list) and marks_list[i] else 1,
                    option1=option1s[i] if i < len(option1s) else '',
                    option2=option2s[i] if i < len(option2s) else '',
                    option3=option3s[i] if i < len(option3s) else '',
                    option4=option4s[i] if i < len(option4s) else '',
                    correct_option=int(correct_options[i]) if i < len(correct_options) and correct_options[i] else None,
                )
        _notify_all_students(
            'contest',
            f'New Contest: {contest.title}',
            f'{request.user.username} created a new contest — "{contest.title}" ({contest.subject.name}, {contest.class_obj.name})',
            link=f'/contests/{contest.pk}/',
            title_bn=f'নতুন Contest: {contest.title}',
            message_bn=f'{request.user.username} একটি নতুন contest তৈরি করেছেন — "{contest.title}" ({contest.subject.name}, {contest.class_obj.name})',
        )
        messages.success(request, 'Contest created!')
        return redirect('contest_detail', pk=contest.pk)
    return render(request, 'core/contest_create.html', {
        'subjects': subjects,
        'classes': classes,
    })


@login_required
def contest_detail(request, pk):
    from .models import (
        Contest, ContestSubmission, ContestRegistration, UserRating,
    )
    from django.utils import timezone
    from django.db.models import F
    try:
        contest = Contest.objects.select_related(
            'subject', 'class_obj', 'created_by'
        ).get(pk=pk)
    except Contest.DoesNotExist:
        messages.warning(request, 'This contest is no longer available (deleted).')
        return redirect('contest_list')

    Contest.objects.filter(pk=pk).update(view_count=F('view_count') + 1)

    now = timezone.now()
    has_submitted = ContestSubmission.objects.filter(
        contest=contest, student=request.user, is_submitted=True,
    ).exists()
    my_registration = ContestRegistration.objects.filter(
        contest=contest, user=request.user,
    ).first()
    is_active = contest.start_time <= now <= contest.end_time

    participant_count = ContestRegistration.objects.filter(contest=contest).count()
    recent_participants = ContestRegistration.objects.filter(
        contest=contest,
    ).select_related('user').order_by('-registered_at')[:8]

    question_count = contest.questions.count()
    my_rating, _ = UserRating.objects.get_or_create(user=request.user)

    leaderboard_preview = None
    if contest.is_past or (not contest.hide_leaderboard_until_end and is_active):
        leaderboard_preview = ContestSubmission.objects.filter(
            contest=contest, is_submitted=True,
            student__profile__role='STUDENT',
        ).select_related('student').order_by(
            '-total_marks', 'duration_taken',
        )[:10]

    return render(request, 'core/contest_detail.html', {
        'contest': contest,
        'has_submitted': has_submitted,
        'is_active': is_active,
        'now': now,
        'my_registration': my_registration,
        'participant_count': participant_count,
        'recent_participants': recent_participants,
        'question_count': question_count,
        'my_rating': my_rating,
        'leaderboard_preview': leaderboard_preview,
    })


@login_required
def contest_join(request, pk):
    from .models import Contest, ContestSubmission
    from django.utils import timezone
    try:
        if request.user.profile.role == 'ADMIN' or request.user.profile.is_superadmin:
            messages.info(request, _L(request, 'Teachers/Admins cannot join contests — view only.', 'Teacher/Admin contest-এ অংশ নিতে পারবেন না — শুধু দেখতে পারবেন।'))
            return redirect('contest_detail', pk=pk)
    except Exception:
        pass
    try:
        contest = Contest.objects.get(pk=pk)
    except Contest.DoesNotExist:
        messages.warning(request, _L(request, 'This contest is no longer available.', 'এই contest আর available নেই।'))
        return redirect('contest_list')
    now = timezone.now()

    if now < contest.start_time:
        messages.error(request, _L(request, 'The contest has not started yet.', 'Contest এখনো শুরু হয়নি।'))
        return redirect('contest_detail', pk=pk)
    if now > contest.end_time:
        messages.error(request, _L(request, 'The contest has ended.', 'Contest শেষ হয়ে গেছে।'))
        return redirect('contest_detail', pk=pk)

    submission, created = ContestSubmission.objects.get_or_create(
        contest=contest,
        student=request.user,
    )
    if submission.is_submitted:
        messages.error(request, _L(request, 'You have already submitted.', 'তুমি আগেই submit করেছ।'))
        return redirect('contest_leaderboard', pk=pk)

    questions = contest.questions.all()
    return render(request, 'core/contest_exam.html', {
        'contest': contest,
        'questions': questions,
        'submission': submission,
    })


@login_required
def contest_submit(request, pk):
    from .models import Contest, ContestSubmission, ContestAnswer
    from django.utils import timezone
    try:
        if request.user.profile.role == 'ADMIN' or request.user.profile.is_superadmin:
            messages.error(request, _L(request, 'Teachers/Admins cannot submit contest answers.', 'Teacher/Admin contest answer submit করতে পারবেন না।'))
            return redirect('contest_detail', pk=pk)
    except Exception:
        pass
    if request.method != 'POST':
        return redirect('contest_detail', pk=pk)

    contest = get_object_or_404(Contest, pk=pk)
    submission = get_object_or_404(ContestSubmission, contest=contest, student=request.user)

    if submission.is_submitted:
        return redirect('contest_leaderboard', pk=pk)

    # Answer rows and the final submission state must land together — a partial
    # set of ContestAnswers with is_submitted=False would block resubmission
    # while scoring only half the paper.
    with transaction.atomic():
        total_marks = 0
        questions = contest.questions.all()

        for q in questions:
            if q.question_type == 'MCQ':
                answer_val = request.POST.get(f'q_{q.pk}')
                mcq_answer = int(answer_val) if answer_val else None
                is_correct = mcq_answer == q.correct_option if mcq_answer else False
                marks_obtained = q.marks if is_correct else 0
                total_marks += marks_obtained
                ContestAnswer.objects.create(
                    submission=submission,
                    question=q,
                    mcq_answer=mcq_answer,
                    is_correct=is_correct,
                    marks_obtained=marks_obtained
                )
            else:
                creative_answer = request.POST.get(f'q_{q.pk}', '')
                ContestAnswer.objects.create(
                    submission=submission,
                    question=q,
                    creative_answer=creative_answer,
                    is_correct=None,
                    marks_obtained=0
                )

        now = timezone.now()
        duration = int((now - submission.started_at).total_seconds())
        submission.submitted_at = now
        submission.total_marks = total_marks
        submission.duration_taken = duration
        submission.time_taken_seconds = duration
        submission.is_submitted = True

        from .models import ContestRegistration
        reg = ContestRegistration.objects.filter(
            contest=contest, user=request.user,
        ).first()
        if reg is not None:
            submission.is_rated_participant = reg.is_rated and contest.is_rated
        else:
            submission.is_rated_participant = contest.is_rated
        submission.save()

    logger.info('Contest submission: contest=%s user=%s marks=%s',
                contest.pk, request.user.username, total_marks)
    messages.success(request, f'Submitted! Your marks: {total_marks}')
    return redirect('contest_result', pk=pk)


@login_required
def contest_result(request, pk):
    from .models import Contest, ContestSubmission, ContestAnswer
    from django.db.models import Sum
    try:
        if request.user.profile.role == 'ADMIN' or request.user.profile.is_superadmin:
            return redirect('contest_leaderboard', pk=pk)
    except Exception:
        pass
    contest = get_object_or_404(Contest, pk=pk)
    submission = get_object_or_404(ContestSubmission, contest=contest, student=request.user, is_submitted=True)
    answers = ContestAnswer.objects.filter(submission=submission).select_related('question')
    max_marks = contest.questions.aggregate(total=Sum('marks'))['total'] or 0
    correct_count = answers.filter(is_correct=True).count()
    percentage = round(submission.total_marks / max_marks * 100, 1) if max_marks > 0 else 0
    all_subs = list(ContestSubmission.objects.filter(
        contest=contest, is_submitted=True,
        student__profile__role='STUDENT',
    ).order_by('-total_marks', 'duration_taken').values_list('student_id', flat=True))
    rank = next((i + 1 for i, uid in enumerate(all_subs) if uid == request.user.id), None)
    return render(request, 'core/contest_result.html', {
        'contest': contest,
        'submission': submission,
        'answers': answers,
        'correct_count': correct_count,
        'max_marks': max_marks,
        'percentage': percentage,
        'rank': rank,
        'total_participants': len(all_subs),
    })


@login_required
def contest_leaderboard(request, pk):
    from .models import Contest, ContestSubmission, UserRating
    from django.db.models import Sum
    from django.utils import timezone

    contest = get_object_or_404(Contest, pk=pk)
    now = timezone.now()
    is_live = contest.start_time <= now <= contest.end_time
    hide = contest.hide_leaderboard_until_end and is_live

    submissions_qs = ContestSubmission.objects.filter(
        contest=contest, is_submitted=True,
        student__profile__role='STUDENT',
    ).select_related('student', 'student__profile').order_by(
        '-total_marks', 'duration_taken',
    )

    submissions = list(submissions_qs) if not hide else []
    user_rating_map = {
        ur.user_id: ur for ur in UserRating.objects.filter(
            user_id__in=[s.student_id for s in submissions]
        )
    }
    for s in submissions:
        s.user_rating_obj = user_rating_map.get(s.student_id)

    my_submission = ContestSubmission.objects.filter(
        contest=contest, student=request.user, is_submitted=True,
    ).first()
    max_marks = contest.questions.aggregate(total=Sum('marks'))['total'] or 0
    podium = submissions[:3]
    rest = submissions[3:]
    return render(request, 'core/contest_leaderboard.html', {
        'contest': contest,
        'submissions': submissions,
        'podium': podium,
        'rest': rest,
        'my_submission': my_submission,
        'now': now,
        'max_marks': max_marks,
        'is_live': is_live,
        'hide_leaderboard': hide,
    })


@admin_required
def contest_stats(request, pk):
    from .models import Contest, ContestSubmission, ContestAnswer
    from django.db.models import Sum, Count, Q
    contest = get_object_or_404(Contest, pk=pk)
    submissions = ContestSubmission.objects.filter(
        contest=contest, is_submitted=True,
        student__profile__role='STUDENT',
    ).select_related('student')
    total_participants = submissions.count()
    max_possible = contest.questions.aggregate(total=Sum('marks'))['total'] or 0

    if total_participants > 0:
        marks_list = list(submissions.values_list('total_marks', flat=True))
        avg_marks = round(sum(marks_list) / len(marks_list), 1)
        highest = max(marks_list)
        lowest = min(marks_list)
        pass_count = sum(1 for m in marks_list if m >= max_possible * 0.5)
        pass_rate = round(pass_count / total_participants * 100, 1)
    else:
        avg_marks = highest = lowest = pass_count = pass_rate = 0

    question_stats = []
    for q in contest.questions.all():
        ans = ContestAnswer.objects.filter(submission__in=submissions, question=q)
        t = ans.count()
        c = ans.filter(is_correct=True).count()
        question_stats.append({
            'question': q,
            'total': t,
            'correct': c,
            'accuracy': round(c / t * 100, 1) if t > 0 else 0,
        })
    question_stats.sort(key=lambda x: x['accuracy'])

    return render(request, 'core/contest_stats.html', {
        'contest': contest,
        'total_participants': total_participants,
        'avg_marks': avg_marks,
        'highest': highest,
        'lowest': lowest,
        'max_possible': max_possible,
        'pass_rate': pass_rate,
        'pass_count': pass_count,
        'question_stats': question_stats,
        'top_submissions': submissions.order_by('-total_marks', 'duration_taken')[:10],
    })


@admin_required
def contest_bank_questions(request):
    from .models import Question
    from django.http import JsonResponse
    subject_id = request.GET.get('subject')
    class_id = request.GET.get('class_obj')
    qs = Question.objects.filter(is_active=True, question_type='MCQ')
    if subject_id:
        qs = qs.filter(subject_id=subject_id)
    if class_id:
        qs = qs.filter(class_obj_id=class_id)
    data = [{'id': q.id, 'text': q.question_text[:120], 'difficulty': q.difficulty,
              'option1': q.option1, 'option2': q.option2, 'option3': q.option3,
              'option4': q.option4, 'correct_option': q.correct_option}
            for q in qs[:60]]
    return JsonResponse({'questions': data})


@admin_required
def contest_delete(request, pk):
    from .models import Contest
    contest = get_object_or_404(Contest, pk=pk)
    if request.method == 'POST':
        contest.delete()
        messages.success(request, 'Contest deleted!')
        return redirect('contest_list')
    return redirect('contest_detail', pk=pk)


@login_required
def profile_view(request):
    from .models import UserProgress, UserRating, UserBadge
    from django.utils import timezone

    profile = request.user.profile
    ctx = {'profile': profile}
    if profile.role == 'ADMIN':
        ctx['all_subjects'] = Subject.objects.filter(is_active=True).order_by('name')

    rating, _ = UserRating.objects.get_or_create(user=request.user)
    progress = UserProgress.objects.filter(user=request.user)
    total = progress.count()
    correct = progress.filter(is_correct=True).count()

    days_left = None
    if profile.plan_expires_at:
        days_left = max((profile.plan_expires_at - timezone.now()).days, 0)

    # Rating band progress (toward next rank)
    thresholds = [0, 800, 1000, 1200, 1400, 1600, 1800]
    r = rating.rating
    lo = max([t for t in thresholds if t <= r] or [0])
    higher = [t for t in thresholds if t > r]
    hi = min(higher) if higher else lo + 200
    band_pct = round((r - lo) / (hi - lo) * 100) if hi > lo else 100

    ctx.update({
        'rating': rating,
        'rank': rating.rank_title,
        'next_rank': rating.next_rank_info,
        'band_pct': band_pct,
        'total_answered': total,
        'total_correct': correct,
        'accuracy': round(correct / total * 100) if total else 0,
        'badges': list(UserBadge.objects.filter(user=request.user).select_related('badge')[:10]),
        'badge_count': UserBadge.objects.filter(user=request.user).count(),
        'days_left': days_left,
        'study_subjects': profile.study_subjects.all(),
    })
    return render(request, 'core/profile.html', ctx)


@login_required
def profile_picture_delete(request):
    if request.method == 'POST':
        profile = request.user.profile
        if profile.profile_picture:
            profile.profile_picture.delete(save=False)
            profile.profile_picture = None
            profile.save(update_fields=['profile_picture'])
            messages.success(request, _L(request, 'Picture removed.', 'ছবি মুছে ফেলা হয়েছে।'))
    return redirect('profile')


@login_required
def profile_update(request):
    if request.method == 'POST':
        user = request.user
        profile = user.profile

        if request.FILES.get('profile_picture'):
            err = _upload_error(request.FILES['profile_picture'], kind='image', max_mb=5)
            if err:
                messages.error(request, _L(request, *err))
                return redirect('profile')
            profile.profile_picture = request.FILES['profile_picture']
            profile.save(update_fields=['profile_picture'])
            messages.success(request, _L(request, 'Profile picture updated!', 'প্রোফাইল ছবি আপডেট হয়েছে!'))
            return redirect('profile')

        if 'update_subjects' in request.POST:
            subject_ids = request.POST.getlist('subjects')
            profile.subjects.set(subject_ids)
            messages.success(request, 'Teaching subjects updated!')
            return redirect('profile')

        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.email = request.POST.get('email', user.email)
        user.save()
        messages.success(request, _L(request, 'Profile updated!', 'প্রোফাইল আপডেট হয়েছে!'))
        return redirect('profile')
    ctx = {'profile': request.user.profile}
    if request.user.profile.role == 'ADMIN':
        ctx['all_subjects'] = Subject.objects.filter(is_active=True).order_by('name')
    return render(request, 'core/profile.html', ctx)


# -------- SYLLABUS --------

def syllabus_list(request):
    from .models import Syllabus
    syllabi = Syllabus.objects.filter(is_active=True).select_related('subject', 'class_obj', 'board')
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    boards = Board.objects.filter(is_active=True)

    subject_filter = request.GET.get('subject')
    class_filter = request.GET.get('class_obj')
    board_filter = request.GET.get('board')

    if subject_filter:
        syllabi = syllabi.filter(subject__slug=subject_filter)
    if class_filter:
        syllabi = syllabi.filter(class_obj__id=class_filter)
    if board_filter:
        syllabi = syllabi.filter(board__id=board_filter)

    return render(request, 'core/syllabus_list.html', {
        'syllabi': syllabi,
        'subjects': subjects,
        'classes': classes,
        'boards': boards,
    })


def syllabus_detail(request, pk):
    from .models import Syllabus
    syllabus = get_object_or_404(Syllabus, pk=pk, is_active=True)
    return render(request, 'core/syllabus_detail.html', {
        'syllabus': syllabus,
    })


@admin_required
def syllabus_add(request):
    from .models import Syllabus
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    boards = Board.objects.filter(is_active=True)
    if request.method == 'POST':
        Syllabus.objects.create(
            subject=get_object_or_404(Subject, pk=request.POST.get('subject')),
            class_obj=get_object_or_404(Class, pk=request.POST.get('class_obj')),
            board=get_object_or_404(Board, pk=request.POST.get('board')),
            content=request.POST.get('content', ''),
            is_active=True
        )
        messages.success(request, 'Syllabus added!')
        return redirect('syllabus_list')
    return render(request, 'core/syllabus_form.html', {
        'subjects': subjects,
        'classes': classes,
        'boards': boards,
        'action': 'Add',
    })


@admin_required
def syllabus_edit(request, pk):
    from .models import Syllabus
    syllabus = get_object_or_404(Syllabus, pk=pk)
    subjects = Subject.objects.filter(is_active=True)
    classes = Class.objects.all()
    boards = Board.objects.filter(is_active=True)
    if request.method == 'POST':
        syllabus.subject = get_object_or_404(Subject, pk=request.POST.get('subject'))
        syllabus.class_obj = get_object_or_404(Class, pk=request.POST.get('class_obj'))
        syllabus.board = get_object_or_404(Board, pk=request.POST.get('board'))
        syllabus.content = request.POST.get('content', '')
        syllabus.save()
        messages.success(request, 'Syllabus updated!')
        return redirect('syllabus_detail', pk=syllabus.pk)
    return render(request, 'core/syllabus_form.html', {
        'syllabus': syllabus,
        'subjects': subjects,
        'classes': classes,
        'boards': boards,
        'action': 'Edit',
    })


@admin_required
def syllabus_delete(request, pk):
    from .models import Syllabus
    syllabus = get_object_or_404(Syllabus, pk=pk)
    if request.method == 'POST':
        syllabus.delete()
        messages.success(request, 'Syllabus deleted!')
    return redirect('syllabus_list')
@superadmin_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from .models import UserProgress, TeacherFeedback, StudyNote, Contest, ContestSubmission, NoteBookmark, ExamPaper, ExamAttempt, NoteRequest, WrittenSolveSubmission

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')

    def style_header(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(len(header) + 5, 15)

    ws1 = wb.active
    ws1.title = 'Users'
    headers = ['Username', 'Email', 'Role', 'Plan', 'Is Superadmin', 'Joined']
    style_header(ws1, headers)
    for profile in UserProfile.objects.select_related('user').all():
        ws1.append([
            profile.user.username,
            profile.user.email,
            profile.role,
            profile.plan,
            'Yes' if profile.is_superadmin else 'No',
            profile.user.date_joined.strftime('%d-%m-%Y %H:%M'),
        ])

    ws2 = wb.create_sheet('Questions')
    headers2 = ['Board', 'Subject', 'Class', 'Year', 'Chapter', 'Question', 'Type', 'Difficulty', 'Correct Option', 'Answer Hint']
    style_header(ws2, headers2)
    for q in Question.objects.select_related('board', 'subject', 'class_obj').filter(is_active=True):
        ws2.append([
            q.board.name, q.subject.name, q.class_obj.name, q.year,
            q.chapter, q.question_text, q.question_type, q.difficulty,
            q.correct_option, q.answer_hint,
        ])

    ws3 = wb.create_sheet('Progress')
    headers3 = ['Student', 'Question', 'Subject', 'Correct', 'Answered At']
    style_header(ws3, headers3)
    for p in UserProgress.objects.select_related('user', 'question', 'question__subject').all():
        ws3.append([
            p.user.username, p.question.question_text[:50],
            p.question.subject.name, 'Yes' if p.is_correct else 'No',
            p.answered_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws4 = wb.create_sheet('Study Notes')
    headers4 = ['Title', 'Subject', 'Class', 'Chapter', 'Created By', 'Created At']
    style_header(ws4, headers4)
    for note in StudyNote.objects.select_related('subject', 'class_obj', 'created_by').all():
        ws4.append([
            note.title, note.subject.name, note.class_obj.name,
            note.chapter, note.created_by.username,
            note.created_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws5 = wb.create_sheet('Contests')
    headers5 = ['Title', 'Subject', 'Class', 'Created By', 'Duration', 'Start', 'End', 'Submissions']
    style_header(ws5, headers5)
    for c in Contest.objects.select_related('subject', 'class_obj', 'created_by').all():
        ws5.append([
            c.title, c.subject.name, c.class_obj.name, c.created_by.username,
            f"{c.duration_minutes} min",
            c.start_time.strftime('%d-%m-%Y %H:%M'),
            c.end_time.strftime('%d-%m-%Y %H:%M'),
            c.submissions.filter(is_submitted=True).count(),
        ])

    ws6 = wb.create_sheet('Contest Results')
    headers6 = ['Contest', 'Student', 'Total Marks', 'Duration (s)', 'Submitted At']
    style_header(ws6, headers6)
    for sub in ContestSubmission.objects.select_related('contest', 'student').filter(is_submitted=True):
        ws6.append([
            sub.contest.title, sub.student.username, sub.total_marks,
            sub.duration_taken,
            sub.submitted_at.strftime('%d-%m-%Y %H:%M') if sub.submitted_at else '',
        ])

    ws7 = wb.create_sheet('Teacher Feedback')
    headers7 = ['Teacher', 'Student', 'Comment', 'Is Read', 'Created At']
    style_header(ws7, headers7)
    for fb in TeacherFeedback.objects.select_related('teacher', 'student').all():
        ws7.append([
            fb.teacher.username, fb.student.username, fb.comment,
            'Yes' if fb.is_read else 'No',
            fb.created_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws8 = wb.create_sheet('Bookmarks')
    headers8 = ['Student', 'Note Title', 'Bookmarked At']
    style_header(ws8, headers8)
    for bm in NoteBookmark.objects.select_related('user', 'note').all():
        ws8.append([
            bm.user.username, bm.note.title,
            bm.created_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws9 = wb.create_sheet('Exam Papers')
    headers9 = ['Title', 'Subject', 'Class', 'Board', 'Year', 'Created By', 'MCQ Count', 'CQ Count', 'Total Attempts', 'Graded', 'Created At']
    style_header(ws9, headers9)
    for paper in ExamPaper.objects.select_related('subject', 'class_obj', 'board', 'created_by').all():
        total_attempts = paper.attempts.count()
        graded = paper.attempts.filter(status='GRADED').count()
        ws9.append([
            paper.title,
            paper.subject.name,
            paper.class_obj.name,
            paper.board.name if paper.board else '',
            paper.year or '',
            paper.created_by.username,
            paper.mcqs.count(),
            paper.cqs.count(),
            total_attempts,
            graded,
            paper.created_at.strftime('%d-%m-%Y %H:%M'),
        ])

    ws10 = wb.create_sheet('Exam Results')
    headers10 = ['Student', 'Exam Paper', 'Subject', 'Status', 'MCQ Score', 'CQ Score', 'Total Score', 'Grade', 'Started At', 'CQ Submitted At']
    style_header(ws10, headers10)
    for attempt in ExamAttempt.objects.select_related('student', 'exam_paper', 'exam_paper__subject').all():
        ws10.append([
            attempt.student.username,
            attempt.exam_paper.title,
            attempt.exam_paper.subject.name,
            attempt.get_status_display(),
            attempt.mcq_score,
            attempt.cq_score if attempt.cq_score is not None else '',
            attempt.total_score if attempt.total_score is not None else '',
            attempt.grade or '',
            attempt.started_at.strftime('%d-%m-%Y %H:%M'),
            attempt.cq_submitted_at.strftime('%d-%m-%Y %H:%M') if attempt.cq_submitted_at else '',
        ])

    ws11 = wb.create_sheet('Note Requests')
    headers11 = ['Student', 'Subject', 'Topic', 'Details', 'Status', 'Requested At', 'Fulfilled At', 'Fulfilled By', 'Fulfilled Note']
    style_header(ws11, headers11)
    for nr in NoteRequest.objects.select_related('student', 'subject', 'fulfilled_by', 'fulfilled_note').all():
        ws11.append([
            nr.student.username,
            nr.subject.name if nr.subject else '',
            nr.topic,
            nr.details,
            nr.status,
            nr.created_at.strftime('%d-%m-%Y %H:%M'),
            nr.fulfilled_at.strftime('%d-%m-%Y %H:%M') if nr.fulfilled_at else '',
            nr.fulfilled_by.username if nr.fulfilled_by else '',
            nr.fulfilled_note.title if nr.fulfilled_note else '',
        ])

    ws12 = wb.create_sheet('Written CQ Submissions')
    headers12 = ['Student', 'Question', 'Subject', 'Chapter', 'Board', 'Year', 'Submitted At']
    style_header(ws12, headers12)
    for ws in WrittenSolveSubmission.objects.select_related('student', 'question', 'question__subject', 'question__board').all():
        ws12.append([
            ws.student.username,
            ws.question.question_text[:60],
            ws.question.subject.name,
            ws.question.chapter or '',
            ws.question.board.name,
            ws.question.year,
            ws.submitted_at.strftime('%d-%m-%Y %H:%M'),
        ])

    from django.utils import timezone
    filename = f"PrepareYourself_Data_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -------- EXAM MODE VIEWS --------

@login_required
def exam_paper_list(request):
    from .models import ExamPaper, ExamAttempt
    papers = ExamPaper.objects.filter(is_active=True).select_related(
        'subject', 'class_obj', 'board'
    ).order_by('-year', 'subject__name')

    attempt_map = {}
    if request.user.is_authenticated:
        for a in ExamAttempt.objects.filter(
            student=request.user, exam_paper__is_active=True
        ).values('exam_paper_id', 'status', 'id'):
            attempt_map[a['exam_paper_id']] = a

    paper_data = []
    for paper in papers:
        attempt = attempt_map.get(paper.id)
        paper_data.append({
            'paper': paper,
            'mcq_count': paper.mcqs.count(),
            'cq_count': paper.cqs.count(),
            'attempt': attempt,
        })

    subjects = ExamPaper.objects.filter(is_active=True).values_list(
        'subject__name', flat=True
    ).distinct().order_by('subject__name')
    classes = ExamPaper.objects.filter(is_active=True).values_list(
        'class_obj__name', flat=True
    ).distinct().order_by('class_obj__numeric_value')

    return render(request, 'core/exam_paper_list.html', {
        'paper_data': paper_data,
        'subjects': subjects,
        'classes': classes,
    })


@login_required
def exam_paper_detail(request, pk):
    from .models import ExamPaper, ExamAttempt
    paper = get_object_or_404(ExamPaper, pk=pk, is_active=True)

    attempt = ExamAttempt.objects.filter(
        exam_paper=paper, student=request.user
    ).first()

    mcqs = paper.mcqs.all()[:30]
    cqs = paper.cqs.all()

    return render(request, 'core/exam_paper_detail.html', {
        'paper': paper,
        'attempt': attempt,
        'mcq_count': mcqs.count(),
        'cq_count': cqs.count(),
        'mcq_sample': mcqs[:3],
    })


@login_required
def preview_exam(request, pk):
    from .models import ExamPaper
    if not _is_exam_staff(request.user):
        messages.error(request, _L(request, 'Only Teachers/Admins can view this page.', 'শুধুমাত্র Teacher/Admin এই page দেখতে পারবেন।'))
        return redirect('exam_paper_detail', pk=pk)
    paper = get_object_or_404(ExamPaper, pk=pk, is_active=True)
    mcqs = paper.mcqs.all()
    cqs = paper.cqs.all()
    return render(request, 'manage/preview_exam.html', {
        'paper': paper,
        'mcqs': mcqs,
        'cqs': cqs,
    })


def _calculate_grade(score, max_marks):
    if not max_marks or max_marks <= 0:
        return '—'
    pct = (score / max_marks) * 100
    if pct >= 80:
        return 'A+'
    elif pct >= 70:
        return 'A'
    elif pct >= 60:
        return 'A-'
    elif pct >= 50:
        return 'B'
    elif pct >= 40:
        return 'C'
    elif pct >= 33:
        return 'D'
    return 'F'


EXAM_MCQ_MAX = 30
EXAM_CQ_MAX = 70
EXAM_TOTAL_MAX = 100


def _exam_max_marks(attempt=None):
    return EXAM_TOTAL_MAX


def _auto_submit_mcq(attempt):
    from django.utils import timezone
    mcqs = attempt.exam_paper.mcqs.all()[:30]
    score = 0
    for mcq in mcqs:
        selected = int(attempt.mcq_answers.get(str(mcq.id), 0))
        if selected == mcq.correct_option:
            score += mcq.marks
    attempt.mcq_score = score
    attempt.mcq_submitted_at = timezone.now()
    if attempt.exam_paper.cqs.exists():
        attempt.status = 'MCQ_DONE'
    else:
        attempt.status = 'GRADED'
        attempt.cq_score = 0
        attempt.total_score = score
        attempt.grade = _calculate_grade(score, _exam_max_marks(attempt))
    attempt.save()


@login_required
def start_exam(request, pk):
    from .models import ExamPaper, ExamAttempt
    try:
        if request.user.profile.role == 'ADMIN' or request.user.profile.is_superadmin:
            messages.error(request, _L(request, 'Teachers/Admins cannot take exams.', 'Teacher/Admin রা পরীক্ষা দিতে পারবেন না।'))
            return redirect('exam_paper_detail', pk=pk)
    except Exception:
        pass
    paper = get_object_or_404(ExamPaper, pk=pk, is_active=True)
    attempt = ExamAttempt.objects.filter(exam_paper=paper, student=request.user).first()

    if attempt:
        if attempt.status == 'MCQ_PHASE':
            if attempt.mcq_seconds_remaining == 0:
                _auto_submit_mcq(attempt)
                return redirect('exam_cq_phase', attempt_id=attempt.id)
            mcqs = list(paper.mcqs.all()[:30])
            return render(request, 'core/exam_mcq_phase.html', {
                'paper': paper,
                'attempt': attempt,
                'mcqs': mcqs,
                'seconds_remaining': attempt.mcq_seconds_remaining,
            })
        if attempt.status in ('MCQ_DONE', 'CQ_PHASE'):
            return redirect('exam_cq_phase', attempt_id=attempt.id)
        return redirect('exam_results', attempt_id=attempt.id)

    from django.utils import timezone
    mcqs = list(paper.mcqs.all()[:30])
    if not mcqs:
        attempt = ExamAttempt.objects.create(
            exam_paper=paper, student=request.user,
            status='MCQ_DONE', mcq_score=0, mcq_submitted_at=timezone.now()
        )
        if not paper.cqs.exists():
            attempt.status = 'GRADED'
            attempt.cq_score = 0
            attempt.total_score = 0
            attempt.grade = _calculate_grade(0, _exam_max_marks(attempt))
            attempt.save()
            return redirect('exam_results', attempt_id=attempt.id)
        return redirect('exam_cq_phase', attempt_id=attempt.id)

    attempt = ExamAttempt.objects.create(exam_paper=paper, student=request.user, status='MCQ_PHASE')
    return render(request, 'core/exam_mcq_phase.html', {
        'paper': paper,
        'attempt': attempt,
        'mcqs': mcqs,
        'seconds_remaining': attempt.mcq_seconds_remaining,
    })


@login_required
def submit_mcq(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    try:
        if request.user.profile.role == 'ADMIN' or request.user.profile.is_superadmin:
            return JsonResponse({'error': 'Not allowed'}, status=403)
    except Exception:
        pass
    try:
        data = json.loads(request.body)
        attempt_id = int(data.get('attempt_id', 0))
        answers = data.get('answers', {})
    except (json.JSONDecodeError, ValueError, TypeError):
        return JsonResponse({'error': 'Invalid data'}, status=400)

    from .models import ExamAttempt
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, student=request.user, status='MCQ_PHASE')
    from django.utils import timezone

    mcqs = attempt.exam_paper.mcqs.all()[:30]
    score = 0
    for mcq in mcqs:
        selected = int(answers.get(str(mcq.id), 0))
        if selected == mcq.correct_option:
            score += mcq.marks

    attempt.mcq_answers = {str(k): int(v) for k, v in answers.items() if str(v).isdigit()}
    attempt.mcq_score = score
    attempt.mcq_submitted_at = timezone.now()
    if attempt.exam_paper.cqs.exists():
        attempt.status = 'MCQ_DONE'
        redirect_url = f'/exam/{attempt.id}/cq/'
    else:
        attempt.status = 'GRADED'
        attempt.cq_score = 0
        attempt.total_score = score
        attempt.grade = _calculate_grade(score, _exam_max_marks(attempt))
        redirect_url = f'/exam/{attempt.id}/results/'
    attempt.save()

    return JsonResponse({
        'success': True,
        'score': score,
        'redirect': redirect_url,
    })


@login_required
def exam_cq_phase(request, attempt_id):
    try:
        if request.user.profile.role == 'ADMIN' or request.user.profile.is_superadmin:
            messages.error(request, _L(request, 'Teachers/Admins cannot take exams.', 'Teacher/Admin রা পরীক্ষা দিতে পারবেন না।'))
            return redirect('exam_paper_list')
    except Exception:
        pass
    from .models import ExamAttempt
    from django.utils import timezone
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, student=request.user)

    if attempt.status == 'MCQ_PHASE':
        return redirect('start_exam', pk=attempt.exam_paper.id)
    if attempt.status in ('CQ_PENDING', 'GRADED'):
        return redirect('exam_results', attempt_id=attempt.id)

    if attempt.status == 'MCQ_DONE':
        attempt.status = 'CQ_PHASE'
        attempt.cq_started_at = timezone.now()
        attempt.save()

    if attempt.cq_seconds_remaining == 0:
        attempt.status = 'CQ_PENDING'
        attempt.cq_submitted_at = timezone.now()
        attempt.save()
        messages.info(request, _L(request, 'CQ time is up. Answers were auto-submitted.', 'CQ সময় শেষ। উত্তর auto-submit হয়েছে।'))
        return redirect('exam_results', attempt_id=attempt.id)

    cqs = attempt.exam_paper.cqs.all()
    existing_subs = {s.cq_question_id: s for s in attempt.cq_submissions.all()}
    selected_ids = set(attempt.selected_cqs)

    return render(request, 'core/exam_cq_phase.html', {
        'attempt': attempt,
        'cqs': cqs,
        'existing_subs': existing_subs,
        'selected_ids': selected_ids,
        'seconds_remaining': attempt.cq_seconds_remaining,
    })


@login_required
def submit_cq(request, attempt_id):
    try:
        if request.user.profile.role == 'ADMIN' or request.user.profile.is_superadmin:
            messages.error(request, _L(request, 'Teachers/Admins cannot take exams.', 'Teacher/Admin রা পরীক্ষা দিতে পারবেন না।'))
            return redirect('exam_paper_list')
    except Exception:
        pass
    if request.method != 'POST':
        return redirect('exam_cq_phase', attempt_id=attempt_id)

    from .models import ExamAttempt, CQQuestion, CQSubmission
    from django.utils import timezone
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, student=request.user)

    if attempt.status in ('CQ_PENDING', 'GRADED'):
        messages.info(request, _L(request, 'This exam has already been submitted.', 'এই পরীক্ষার উত্তর আগেই জমা হয়েছে।'))
        return redirect('exam_results', attempt_id=attempt.id)
    if attempt.status == 'MCQ_PHASE':
        messages.error(request, _L(request, 'The MCQ phase is not finished yet.', 'এখনো MCQ পর্ব শেষ হয়নি।'))
        return redirect('start_exam', pk=attempt.exam_paper.id)
    if attempt.status not in ('CQ_PHASE', 'MCQ_DONE'):
        messages.error(request, _L(request, 'CQ cannot be submitted right now.', 'এই মুহূর্তে CQ submit করা যাচ্ছে না।'))
        return redirect('exam_results', attempt_id=attempt.id)

    selected_cq_ids = []
    for val in request.POST.getlist('selected_cqs'):
        try:
            selected_cq_ids.append(int(val))
        except ValueError:
            pass
    selected_cq_ids = selected_cq_ids[:7]

    for uploaded in request.FILES.values():
        err = _upload_error(uploaded, kind='image', max_mb=10)
        if err:
            messages.error(request, _L(request, *err))
            return redirect('exam_cq_phase', attempt_id=attempt.id)

    # CQ answer rows and the attempt's CQ_PENDING status must commit together,
    # otherwise a mid-request failure leaves answers saved but the attempt
    # still in CQ_PHASE (or vice versa).
    with transaction.atomic():
        for cq_id in selected_cq_ids:
            try:
                cq = CQQuestion.objects.get(id=cq_id, exam_paper=attempt.exam_paper)
            except CQQuestion.DoesNotExist:
                continue
            sub, _ = CQSubmission.objects.get_or_create(attempt=attempt, cq_question=cq)
            if f'photo_{cq_id}' in request.FILES:
                sub.photo = request.FILES[f'photo_{cq_id}']
            for part in ('a', 'b', 'c', 'd'):
                key = f'photo_{cq_id}_{part}'
                if key in request.FILES:
                    setattr(sub, f'photo_{part}', request.FILES[key])
            sub.save()

        attempt.selected_cqs = selected_cq_ids
        attempt.status = 'CQ_PENDING'
        attempt.cq_submitted_at = timezone.now()
        attempt.save()

    logger.info('CQ submitted: attempt=%s user=%s paper=%s',
                attempt.id, request.user.username, attempt.exam_paper.title)

    # In-app notification — one per (teacher, exam paper) while unread, subject-filtered
    from .models import Notification, UserProfile, ExamAttempt as _EA
    paper_title = attempt.exam_paper.title
    paper_subject = attempt.exam_paper.subject
    grade_link = '/manage/grade-queue/'

    subject_teachers = UserProfile.objects.filter(
        role='ADMIN', is_approved=True, subjects=paper_subject
    ).select_related('user')
    # Fallback: if no teacher has subjects assigned yet, notify all approved teachers
    if not subject_teachers.exists():
        subject_teachers = UserProfile.objects.filter(
            role='ADMIN', is_approved=True
        ).select_related('user')
    superadmin_profiles = UserProfile.objects.filter(is_superadmin=True).select_related('user')
    notify_users = {p.user for p in subject_teachers} | {p.user for p in superadmin_profiles}

    pending_count = _EA.objects.filter(exam_paper=attempt.exam_paper, status='CQ_PENDING').count()

    for teacher in notify_users:
        # Skip if this teacher already has an unread notification for this paper
        already = Notification.objects.filter(
            recipient=teacher, notif_type='exam', is_read=False,
            title__contains=paper_title,
        ).exists()
        if already:
            continue
        Notification.objects.create(
            recipient=teacher,
            notif_type='exam',
            title=f'CQ submissions waiting — {paper_title}',
            title_bn=f'CQ জমা অপেক্ষায় — {paper_title}',
            message=f'{pending_count} student(s) submitted CQ answers for "{paper_title}". Check the grade queue.',
            message_bn=f'"{paper_title}" পরীক্ষায় {pending_count} জন CQ উত্তর জমা দিয়েছে। গ্রেড কিউ দেখুন।',
            link=grade_link,
        )

    messages.success(request, _L(request, 'Your CQ answers were submitted successfully. Please wait for evaluation by the teacher.', 'আপনার CQ উত্তর সফলভাবে জমা হয়েছে। Teacher এর মূল্যায়নের জন্য অপেক্ষা করুন।'))
    return redirect('exam_results', attempt_id=attempt.id)


@login_required
def exam_results(request, attempt_id):
    try:
        if request.user.profile.role == 'ADMIN' or request.user.profile.is_superadmin:
            messages.error(request, _L(request, 'Teachers/Admins cannot take exams.', 'Teacher/Admin রা পরীক্ষা দিতে পারবেন না।'))
            return redirect('exam_paper_list')
    except Exception:
        pass
    from .models import ExamAttempt
    attempt = get_object_or_404(ExamAttempt, id=attempt_id, student=request.user)
    cq_submissions = attempt.cq_submissions.select_related('cq_question').all()
    max_marks = _exam_max_marks(attempt)
    score = attempt.total_score or 0
    percentage = round((score / max_marks) * 100, 1) if max_marks else 0
    live_grade = _calculate_grade(score, max_marks) if attempt.status == 'GRADED' else attempt.grade
    if attempt.status == 'GRADED' and live_grade != attempt.grade:
        attempt.grade = live_grade
        attempt.save(update_fields=['grade'])
    return render(request, 'core/exam_results.html', {
        'attempt': attempt,
        'cq_submissions': cq_submissions,
        'max_marks': max_marks,
        'percentage': percentage,
        'live_grade': live_grade,
    })


def _is_exam_staff(user):
    if not user.is_authenticated:
        return False
    if user.is_staff or user.is_superuser:
        return True
    try:
        p = user.profile
        return p.is_superadmin or p.role == 'ADMIN'
    except Exception:
        return False


def _notify_all_students(notif_type, title, message, link='', title_bn='', message_bn=''):
    from .models import Notification, UserProfile
    student_ids = UserProfile.objects.filter(role='STUDENT').values_list('user_id', flat=True)
    Notification.objects.bulk_create([
        Notification(
            recipient_id=uid, notif_type=notif_type,
            title=title, message=message,
            title_bn=title_bn, message_bn=message_bn,
            link=link,
        )
        for uid in student_ids
    ])


@login_required
def create_exam_paper(request):
    from .models import ExamPaper, ExamPaperMCQ, CQQuestion
    if not _is_exam_staff(request.user):
        messages.error(request, _L(request, 'Only Teachers/Staff can access this page.', 'Teacher/Staff শুধু এই page access করতে পারবে।'))
        return redirect('home')

    subjects = Subject.objects.filter(is_active=True).order_by('name')
    classes = Class.objects.all().order_by('numeric_value')
    boards = Board.objects.filter(is_active=True).order_by('name')
    years = list(range(CURRENT_YEAR, CURRENT_YEAR - 10, -1))

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        subject_id = request.POST.get('subject', '').strip()
        class_id = request.POST.get('class_obj', '').strip()
        board_id = request.POST.get('board', '').strip()
        year = request.POST.get('year', '').strip()

        errors = []
        if not title:
            errors.append(_L(request, 'Enter a title.', 'Title দিন।'))
        if not subject_id:
            errors.append(_L(request, 'Select a subject.', 'Subject বেছে নিন।'))
        if not class_id:
            errors.append(_L(request, 'Select a class.', 'Class বেছে নিন।'))

        # Duplicate check: Board + Year combination must be unique per subject/class
        if board_id and year and subject_id and class_id:
            try:
                year_int = int(year)
            except ValueError:
                year_int = None
            if year_int:
                existing = ExamPaper.objects.filter(
                    board_id=board_id, year=year_int,
                    subject_id=subject_id, class_obj_id=class_id,
                ).select_related('board', 'subject', 'class_obj', 'created_by').first()
                if existing:
                    creator = existing.created_by.get_full_name() or existing.created_by.username
                    errors.append(
                        f'এই Board ({existing.board.name}), Year ({year_int}), '
                        f'Subject ({existing.subject.name}) এবং Class ({existing.class_obj.name}) এর paper '
                        f'ইতিমধ্যে আছে: "{existing.title}" — তৈরি করেছেন {creator}.'
                    )

        mcq_texts = request.POST.getlist('mcq_question_text')
        cq_texts = request.POST.getlist('cq_question_text')
        if not any(t.strip() for t in mcq_texts) and not any(t.strip() for t in cq_texts):
            errors.append(_L(request, 'Add at least 1 MCQ or 1 CQ question.', 'কমপক্ষে ১টি MCQ অথবা ১টি CQ প্রশ্ন দিন।'))

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'manage/create_exam_paper.html', {
                'subjects': subjects, 'classes': classes,
                'boards': boards, 'years': years, 'post': request.POST,
            })

        paper = ExamPaper.objects.create(
            title=title,
            subject_id=subject_id,
            class_obj_id=class_id,
            board_id=board_id if board_id else None,
            year=int(year) if year else None,
            created_by=request.user,
            is_active=True,
        )

        # MCQs
        mcq_opt1 = request.POST.getlist('mcq_option1')
        mcq_opt2 = request.POST.getlist('mcq_option2')
        mcq_opt3 = request.POST.getlist('mcq_option3')
        mcq_opt4 = request.POST.getlist('mcq_option4')
        mcq_correct = request.POST.getlist('mcq_correct_option')
        mcq_marks = request.POST.getlist('mcq_marks')

        order = 0
        for i, text in enumerate(mcq_texts):
            text = text.strip()
            if not text:
                continue
            try:
                correct = int(mcq_correct[i]) if i < len(mcq_correct) else 1
                marks = int(mcq_marks[i]) if i < len(mcq_marks) else 1
            except (ValueError, IndexError):
                correct, marks = 1, 1
            ExamPaperMCQ.objects.create(
                exam_paper=paper,
                question_text=text,
                option1=mcq_opt1[i] if i < len(mcq_opt1) else '',
                option2=mcq_opt2[i] if i < len(mcq_opt2) else '',
                option3=mcq_opt3[i] if i < len(mcq_opt3) else '',
                option4=mcq_opt4[i] if i < len(mcq_opt4) else '',
                correct_option=correct,
                marks=marks,
                order=order,
            )
            order += 1

        # CQs
        cq_part_a = request.POST.getlist('cq_part_a')
        cq_part_b = request.POST.getlist('cq_part_b')
        cq_part_c = request.POST.getlist('cq_part_c')
        cq_part_d = request.POST.getlist('cq_part_d')
        cq_marks_a = request.POST.getlist('cq_marks_a')
        cq_marks_b = request.POST.getlist('cq_marks_b')
        cq_marks_c = request.POST.getlist('cq_marks_c')
        cq_marks_d = request.POST.getlist('cq_marks_d')

        order = 0
        for i, text in enumerate(cq_texts):
            text = text.strip()
            if not text:
                continue
            def _m(lst, idx, default):
                try:
                    return max(0, int(lst[idx])) if idx < len(lst) else default
                except ValueError:
                    return default
            CQQuestion.objects.create(
                exam_paper=paper,
                question_text=text,
                part_a=cq_part_a[i] if i < len(cq_part_a) else '',
                part_b=cq_part_b[i] if i < len(cq_part_b) else '',
                part_c=cq_part_c[i] if i < len(cq_part_c) else '',
                part_d=cq_part_d[i] if i < len(cq_part_d) else '',
                marks_a=_m(cq_marks_a, i, 1),
                marks_b=_m(cq_marks_b, i, 2),
                marks_c=_m(cq_marks_c, i, 3),
                marks_d=_m(cq_marks_d, i, 4),
                order=order,
            )
            order += 1

        _notify_all_students(
            'exam',
            f'New Exam Paper: {paper.title}',
            f'{request.user.username} uploaded a new exam paper — "{paper.title}" ({paper.subject.name}, {paper.class_obj.name})',
            link=f'/exam-papers/{paper.pk}/',
            title_bn=f'নতুন Exam Paper: {paper.title}',
            message_bn=f'{request.user.username} একটি নতুন exam paper আপলোড করেছেন — "{paper.title}" ({paper.subject.name}, {paper.class_obj.name})',
        )
        messages.success(
            request,
            f'"{paper.title}" তৈরি হয়েছে — {paper.mcqs.count()}টি MCQ, {paper.cqs.count()}টি CQ।'
        )
        return redirect('exam_paper_detail', pk=paper.pk)

    return render(request, 'manage/create_exam_paper.html', {
        'subjects': subjects,
        'classes': classes,
        'boards': boards,
        'years': years,
    })


@login_required
def edit_exam_paper(request, pk):
    import json as _json
    from .models import ExamPaper, ExamPaperMCQ, CQQuestion, ExamAttempt
    if not _is_exam_staff(request.user):
        messages.error(request, _L(request, 'Only Teachers/Staff can access this page.', 'Teacher/Staff শুধু এই page access করতে পারবে।'))
        return redirect('home')

    paper = get_object_or_404(ExamPaper, pk=pk)
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    classes = Class.objects.all().order_by('numeric_value')
    boards = Board.objects.filter(is_active=True).order_by('name')
    years = list(range(CURRENT_YEAR, CURRENT_YEAR - 10, -1))
    active_attempt_count = ExamAttempt.objects.filter(
        exam_paper=paper
    ).exclude(status='GRADED').count()

    def _get_existing():
        mcqs = list(paper.mcqs.order_by('order').values(
            'question_text', 'option1', 'option2', 'option3', 'option4',
            'correct_option', 'marks'
        ))
        cqs = list(paper.cqs.order_by('order').values(
            'question_text', 'part_a', 'part_b', 'part_c', 'part_d',
            'marks_a', 'marks_b', 'marks_c', 'marks_d'
        ))
        return _json.dumps(mcqs), _json.dumps(cqs)

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        subject_id = request.POST.get('subject', '').strip()
        class_id = request.POST.get('class_obj', '').strip()
        board_id = request.POST.get('board', '').strip()
        year = request.POST.get('year', '').strip()

        errors = []
        if not title: errors.append(_L(request, 'Enter a title.', 'Title দিন।'))
        if not subject_id: errors.append(_L(request, 'Select a subject.', 'Subject বেছে নিন।'))
        if not class_id: errors.append(_L(request, 'Select a class.', 'Class বেছে নিন।'))

        # Duplicate check: Board + Year combination must be unique (excluding self)
        if board_id and year and subject_id and class_id:
            try:
                year_int = int(year)
            except ValueError:
                year_int = None
            if year_int:
                existing = ExamPaper.objects.filter(
                    board_id=board_id, year=year_int,
                    subject_id=subject_id, class_obj_id=class_id,
                ).exclude(pk=paper.pk).select_related(
                    'board', 'subject', 'class_obj', 'created_by'
                ).first()
                if existing:
                    creator = existing.created_by.get_full_name() or existing.created_by.username
                    errors.append(
                        f'এই Board ({existing.board.name}), Year ({year_int}), '
                        f'Subject ({existing.subject.name}) এবং Class ({existing.class_obj.name}) এর paper '
                        f'ইতিমধ্যে আছে: "{existing.title}" — তৈরি করেছেন {creator}.'
                    )

        mcq_texts = request.POST.getlist('mcq_question_text')
        cq_texts = request.POST.getlist('cq_question_text')
        if not any(t.strip() for t in mcq_texts) and not any(t.strip() for t in cq_texts):
            errors.append(_L(request, 'Add at least 1 MCQ or 1 CQ question.', 'কমপক্ষে ১টি MCQ অথবা ১টি CQ প্রশ্ন দিন।'))

        if errors:
            for err in errors:
                messages.error(request, err)
            mcqs_json, cqs_json = _get_existing()
            return render(request, 'manage/edit_exam_paper.html', {
                'paper': paper, 'subjects': subjects, 'classes': classes,
                'boards': boards, 'years': years,
                'existing_mcqs_json': mcqs_json, 'existing_cqs_json': cqs_json,
                'active_attempt_count': active_attempt_count,
            })

        paper.title = title
        paper.subject_id = subject_id
        paper.class_obj_id = class_id
        paper.board_id = board_id if board_id else None
        paper.year = int(year) if year else None
        paper.save()

        paper.mcqs.all().delete()
        mcq_opt1 = request.POST.getlist('mcq_option1')
        mcq_opt2 = request.POST.getlist('mcq_option2')
        mcq_opt3 = request.POST.getlist('mcq_option3')
        mcq_opt4 = request.POST.getlist('mcq_option4')
        mcq_correct = request.POST.getlist('mcq_correct_option')
        mcq_marks_list = request.POST.getlist('mcq_marks')
        order = 0
        for i, text in enumerate(mcq_texts):
            text = text.strip()
            if not text:
                continue
            try:
                correct = int(mcq_correct[i]) if i < len(mcq_correct) else 1
                marks = int(mcq_marks_list[i]) if i < len(mcq_marks_list) else 1
            except (ValueError, IndexError):
                correct, marks = 1, 1
            ExamPaperMCQ.objects.create(
                exam_paper=paper,
                question_text=text,
                option1=mcq_opt1[i] if i < len(mcq_opt1) else '',
                option2=mcq_opt2[i] if i < len(mcq_opt2) else '',
                option3=mcq_opt3[i] if i < len(mcq_opt3) else '',
                option4=mcq_opt4[i] if i < len(mcq_opt4) else '',
                correct_option=correct,
                marks=marks,
                order=order,
            )
            order += 1

        paper.cqs.all().delete()
        cq_part_a = request.POST.getlist('cq_part_a')
        cq_part_b = request.POST.getlist('cq_part_b')
        cq_part_c = request.POST.getlist('cq_part_c')
        cq_part_d = request.POST.getlist('cq_part_d')
        cq_marks_a = request.POST.getlist('cq_marks_a')
        cq_marks_b = request.POST.getlist('cq_marks_b')
        cq_marks_c = request.POST.getlist('cq_marks_c')
        cq_marks_d = request.POST.getlist('cq_marks_d')
        order = 0
        for i, text in enumerate(cq_texts):
            text = text.strip()
            if not text:
                continue
            def _m(lst, idx, default):
                try:
                    return max(0, int(lst[idx])) if idx < len(lst) else default
                except ValueError:
                    return default
            CQQuestion.objects.create(
                exam_paper=paper,
                question_text=text,
                part_a=cq_part_a[i] if i < len(cq_part_a) else '',
                part_b=cq_part_b[i] if i < len(cq_part_b) else '',
                part_c=cq_part_c[i] if i < len(cq_part_c) else '',
                part_d=cq_part_d[i] if i < len(cq_part_d) else '',
                marks_a=_m(cq_marks_a, i, 1),
                marks_b=_m(cq_marks_b, i, 2),
                marks_c=_m(cq_marks_c, i, 3),
                marks_d=_m(cq_marks_d, i, 4),
                order=order,
            )
            order += 1

        messages.success(
            request,
            f'"{paper.title}" আপডেট হয়েছে — {paper.mcqs.count()}টি MCQ, {paper.cqs.count()}টি CQ।'
        )
        return redirect('exam_paper_detail', pk=paper.pk)

    mcqs_json, cqs_json = _get_existing()
    return render(request, 'manage/edit_exam_paper.html', {
        'paper': paper,
        'subjects': subjects,
        'classes': classes,
        'boards': boards,
        'years': years,
        'existing_mcqs_json': mcqs_json,
        'existing_cqs_json': cqs_json,
        'active_attempt_count': active_attempt_count,
    })


@login_required
def delete_exam_paper(request, pk):
    from .models import ExamPaper, ExamAttempt
    if not _is_exam_staff(request.user):
        messages.error(request, _L(request, 'Only Teachers/Staff can access this page.', 'Teacher/Staff শুধু এই page access করতে পারবে।'))
        return redirect('home')

    if request.method != 'POST':
        return redirect('exam_paper_detail', pk=pk)

    paper = get_object_or_404(ExamPaper, pk=pk)
    active_attempts = ExamAttempt.objects.filter(
        exam_paper=paper,
        student__profile__role='STUDENT',
    ).exclude(status='GRADED').count()

    if active_attempts > 0:
        messages.error(
            request,
            f'{active_attempts}জন student এই exam-এ active আছে। Delete করা সম্ভব নয়।'
        )
        return redirect('exam_paper_detail', pk=pk)

    paper_title = paper.title
    paper.is_active = False
    paper.save()
    messages.success(request, _L(request, f'"{paper_title}" deleted successfully.', f'"{paper_title}" সফলভাবে মুছে ফেলা হয়েছে।'))
    return redirect('exam_paper_list')


def _parse_exam_questions(text):
    import re

    # Options: English a/b/c/d (any case) OR Bengali ক/খ/গ/ঘ, any bracket/dot/space delimiter
    opt_re = re.compile(r'^[(（]?([aAbBcCdDকখগঘ])[)）\.\s]+(.*)', re.UNICODE)
    # Numbered question prefix (Arabic or Bengali numerals)
    q_num_re = re.compile(r'^[০-৯0-9]{1,2}[।\.)\s]+(.*)')

    OPT_IDX = {
        'a': 1, 'A': 1, 'ক': 1,
        'b': 2, 'B': 2, 'খ': 2,
        'c': 3, 'C': 3, 'গ': 3,
        'd': 4, 'D': 4, 'ঘ': 4,
    }

    mcqs = []
    cqs = []

    def flush(q_parts, opts):
        if not opts:
            # Fallback: no letter prefixes — first line = question, next 4 = options
            if len(q_parts) >= 3:
                q_text = q_parts[0]
                options = q_parts[1:5]
                if len(options) == 4:
                    mcqs.append({
                        'question_text': q_text,
                        'option1': options[0], 'option2': options[1],
                        'option3': options[2], 'option4': options[3],
                        'correct_option': 1, 'marks': 1,
                    })
            return
        q_text = ' '.join(q_parts).strip()
        o1, o2, o3, o4 = opts.get(1,''), opts.get(2,''), opts.get(3,''), opts.get(4,'')
        if o1 and o2 and o3 and o4:
            mcqs.append({
                'question_text': q_text,
                'option1': o1, 'option2': o2, 'option3': o3, 'option4': o4,
                'correct_option': 1, 'marks': 1,
            })
        elif o1 or o2:
            cqs.append({
                'question_text': q_text,
                'part_a': o1, 'part_b': o2, 'part_c': o3, 'part_d': o4,
                'marks_a': 1, 'marks_b': 2, 'marks_c': 3, 'marks_d': 4,
            })

    q_parts = []
    opts = {}
    in_options = False

    for raw in text.splitlines():
        line = raw.strip()

        if not line:
            # Blank line ends the current question block
            if opts:
                flush(q_parts, opts)
                q_parts, opts, in_options = [], {}, False
            continue

        om = opt_re.match(line)
        if om:
            idx = OPT_IDX.get(om.group(1))
            if idx:
                opts[idx] = om.group(2).strip()
                in_options = True
            continue

        # Non-option text line
        qm = q_num_re.match(line)
        if qm:
            # Numbered question — flush previous block first
            if opts or in_options:
                flush(q_parts, opts)
                q_parts, opts, in_options = [], {}, False
            q_parts.append(qm.group(1).strip())
        else:
            if in_options and len(opts) < 4:
                # Already collecting options but this line has no prefix —
                # treat it as the next sequential option
                opts[len(opts) + 1] = line
            elif in_options:
                # All 4 options filled, plain text = new question block
                flush(q_parts, opts)
                q_parts, opts, in_options = [line], {}, False
            else:
                q_parts.append(line)

    flush(q_parts, opts)
    return mcqs, cqs


@login_required
def parse_exam_text(request):
    """Receives raw OCR text (from browser-side Tesseract.js) and parses MCQ/CQ structure."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if not _is_exam_staff(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    text = request.POST.get('text', '').strip()
    if not text:
        return JsonResponse({'error': _L(request, 'Please provide text.', 'Text দিন।')}, status=400)

    mcqs, cqs = _parse_exam_questions(text)
    return JsonResponse({
        'success': True,
        'data': {'mcqs': mcqs, 'cqs': cqs},
    })


@login_required
def extract_text_from_image(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)
    if not _is_exam_staff(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    image_file = request.FILES.get('image')
    if not image_file:
        return JsonResponse({'error': _L(request, 'Please provide an image.', 'ছবি দিন।')}, status=400)

    try:
        text = ai_svc.gemini_extract_text(
            image_file.read(),
            mime_type=image_file.content_type or 'image/jpeg',
        )
        return JsonResponse({'success': True, 'text': text})
    except ai_svc.AIServiceError as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def grade_cq_submission(request, attempt_id):
    from .models import ExamAttempt, CQSubmission
    if not _is_exam_staff(request.user):
        messages.error(request, _L(request, 'Only Teachers/Staff can access this page.', 'শুধু Teacher/Staff এই page access করতে পারবে।'))
        return redirect('home')

    attempt = get_object_or_404(
        ExamAttempt.objects.filter(status__in=['CQ_PENDING', 'GRADED']),
        id=attempt_id,
    )

    is_super = getattr(getattr(request.user, 'profile', None), 'is_superadmin', False)
    if attempt.status == 'CQ_PENDING' and attempt.assigned_teacher != request.user and not is_super:
        messages.error(request, _L(request, 'Claim this answer script before grading it.', 'এই answer script grade করতে হলে আগে Claim করুন।'))
        return redirect('manage_grade_list')

    cq_submissions = attempt.cq_submissions.select_related('cq_question').all()
    is_regrade = (attempt.status == 'GRADED')

    if request.method == 'POST':
        from django.utils import timezone
        # Per-question marks and the attempt's final score/grade must commit
        # together — partially saved marks with an unGRADED attempt would let
        # the script be re-claimed and double-graded.
        with transaction.atomic():
            total_cq = 0
            for sub in cq_submissions:
                q = sub.cq_question
                cq_total = 0
                for part, max_marks in (('a', q.marks_a), ('b', q.marks_b),
                                        ('c', q.marks_c), ('d', q.marks_d)):
                    raw = request.POST.get(f'marks_{sub.id}_{part}', '').strip()
                    try:
                        m = int(raw) if raw else 0
                    except ValueError:
                        m = 0
                    m = max(0, min(m, max_marks))
                    setattr(sub, f'marks_{part}', m)
                    comment = request.POST.get(f'comment_{sub.id}_{part}', '').strip()
                    # Clear comment if student got full marks for this part
                    setattr(sub, f'comment_{part}', '' if m >= max_marks else comment)
                    cq_total += m
                sub.marks_given = cq_total
                sub.save()
                total_cq += cq_total

            attempt.cq_score = total_cq
            attempt.total_score = attempt.mcq_score + total_cq
            attempt.grade = _calculate_grade(attempt.total_score, _exam_max_marks(attempt))
            attempt.status = 'GRADED'
            attempt.graded_by = request.user
            attempt.graded_at = timezone.now()
            attempt.save()

        logger.info('CQ graded: attempt=%s student=%s grader=%s score=%s grade=%s%s',
                    attempt.id, attempt.student.username, request.user.username,
                    attempt.total_score, attempt.grade, ' (regrade)' if is_regrade else '')

        from .models import Notification
        from django.urls import reverse
        if is_regrade:
            n_title = f'Result updated: {attempt.exam_paper.title}'
            n_title_bn = f'ফলাফল আপডেট: {attempt.exam_paper.title}'
            n_msg = f'Your result has been updated. New score: {attempt.total_score}, Grade: {attempt.grade}.'
            n_msg_bn = f'তোমার ফলাফল আপডেট হয়েছে। নতুন নম্বর: {attempt.total_score}, গ্রেড: {attempt.grade}।'
        else:
            n_title = f'Result published: {attempt.exam_paper.title}'
            n_title_bn = f'ফলাফল প্রকাশিত: {attempt.exam_paper.title}'
            n_msg = f'Your exam has been graded. Score: {attempt.total_score}, Grade: {attempt.grade}.'
            n_msg_bn = f'তোমার পরীক্ষার ফলাফল প্রকাশিত হয়েছে। নম্বর: {attempt.total_score}, গ্রেড: {attempt.grade}।'
        Notification.objects.create(
            recipient=attempt.student,
            notif_type='exam',
            title=n_title,
            message=n_msg,
            title_bn=n_title_bn,
            message_bn=n_msg_bn,
            link=reverse('exam_results', args=[attempt.id]),
        )

        messages.success(request, _L(request, f'{"Re-grading" if is_regrade else "Grading"} complete. Grade: {attempt.grade}', f'{"Re-grading" if is_regrade else "Grading"} সম্পন্ন। Grade: {attempt.grade}'))
        return redirect('manage_grade_list')

    return render(request, 'manage/grade_cq_submission.html', {
        'attempt': attempt,
        'cq_submissions': cq_submissions,
        'is_regrade': is_regrade,
    })


@login_required
def manage_grade_list(request):
    if not _is_exam_staff(request.user):
        messages.error(request, _L(request, 'Only Teachers/Staff can access this page.', 'শুধু Teacher/Staff এই page access করতে পারবে।'))
        return redirect('home')

    from .models import ExamAttempt
    from django.utils import timezone
    from datetime import timedelta

    is_super = getattr(getattr(request.user, 'profile', None), 'is_superadmin', False)
    try:
        my_subjects = list(request.user.profile.subjects.values_list('id', flat=True))
    except Exception:
        my_subjects = []

    qs_pending = ExamAttempt.objects.filter(status='CQ_PENDING').select_related(
        'student', 'exam_paper', 'exam_paper__subject', 'assigned_teacher'
    ).order_by('cq_submitted_at')

    # Unclaimed: filtered by teacher's current subject
    unclaimed_qs = qs_pending.filter(assigned_teacher=None)
    if not is_super and my_subjects:
        unclaimed_qs = unclaimed_qs.filter(exam_paper__subject__id__in=my_subjects)

    # My queue: always show everything this teacher claimed, even if they changed subject
    my_queue = qs_pending.filter(assigned_teacher=request.user)

    # Graded by me: all attempts I've graded (most recent first)
    graded_by_me = ExamAttempt.objects.filter(
        status='GRADED', graded_by=request.user
    ).select_related(
        'student', 'exam_paper', 'exam_paper__subject'
    ).order_by('-graded_at')[:100]

    total_pending = unclaimed_qs.count() + my_queue.count()
    unclaimed = unclaimed_qs

    cutoff = timezone.now() - timedelta(hours=24)
    for qs in [unclaimed, my_queue]:
        for attempt in qs:
            attempt.is_urgent = bool(attempt.cq_submitted_at and attempt.cq_submitted_at < cutoff)

    return render(request, 'manage/grade_queue.html', {
        'unclaimed': unclaimed,
        'my_queue': my_queue,
        'graded_by_me': graded_by_me,
        'total_pending': total_pending,
    })


@login_required
def claim_cq_attempt(request, attempt_id):
    from .models import ExamAttempt
    if not _is_exam_staff(request.user):
        return redirect('home')
    if request.method == 'POST':
        attempt = get_object_or_404(ExamAttempt, id=attempt_id, status='CQ_PENDING')
        if attempt.assigned_teacher is None:
            from django.utils import timezone
            attempt.assigned_teacher = request.user
            attempt.claimed_at = timezone.now()
            attempt.save()
    return redirect('manage_grade_list')


@login_required
def release_cq_attempt(request, attempt_id):
    from .models import ExamAttempt
    if not _is_exam_staff(request.user):
        return redirect('home')
    if request.method == 'POST':
        attempt = get_object_or_404(ExamAttempt, id=attempt_id, status='CQ_PENDING')
        is_super = getattr(getattr(request.user, 'profile', None), 'is_superadmin', False)
        if attempt.assigned_teacher == request.user or is_super:
            attempt.assigned_teacher = None
            attempt.claimed_at = None
            attempt.save()
    return redirect('manage_grade_list')


# -------- CONTEST: REGISTRATION, RATINGS, BADGES, COINS --------

@login_required
def contest_register(request, pk):
    """POST: register the current user for a contest.

    Honors entry_requirement, registration_deadline, max_participants, and
    is_rated/allow_unrated_join. Awards early_bird bonus where applicable.
    """
    from .models import Contest, ContestRegistration
    from .services import coins as coin_svc
    from .services import badges as badge_svc
    from django.utils import timezone

    if request.method != 'POST':
        return redirect('contest_detail', pk=pk)

    contest = get_object_or_404(Contest, pk=pk)
    now = timezone.now()

    try:
        profile = request.user.profile
        if profile.role == 'ADMIN' or profile.is_superadmin:
            messages.error(request, 'Teachers/admins cannot register for contests.')
            return redirect('contest_detail', pk=pk)
    except Exception:
        pass

    if contest.registration_deadline and now > contest.registration_deadline:
        messages.error(request, 'Registration deadline has passed.')
        return redirect('contest_detail', pk=pk)
    if now > contest.end_time:
        messages.error(request, 'This contest has already ended.')
        return redirect('contest_detail', pk=pk)

    req = contest.entry_requirement
    if req == 'premium' and not getattr(profile, 'is_premium', False):
        messages.error(request, 'Premium membership required for this contest.')
        return redirect('contest_detail', pk=pk)
    class_map = {'class_9': '9', 'class_10': '10', 'class_11': '11', 'class_12': '12'}
    if req in class_map:
        cls_name = (contest.class_obj.name or '').strip()
        if class_map[req] not in cls_name:
            messages.error(request, f'This contest is only for {contest.class_obj.name}.')
            return redirect('contest_detail', pk=pk)

    if contest.max_participants:
        reg_count = ContestRegistration.objects.filter(contest=contest).count()
        if reg_count >= contest.max_participants:
            existing = ContestRegistration.objects.filter(
                contest=contest, user=request.user,
            ).first()
            if not existing:
                messages.error(request, 'Contest is full.')
                return redirect('contest_detail', pk=pk)

    want_rated = (request.POST.get('is_rated', '1') == '1')
    if not contest.is_rated:
        want_rated = False
    if not contest.allow_unrated_join and contest.is_rated:
        want_rated = True

    is_first = not contest.registrations.filter(user=request.user).exists()
    is_early = (now - contest.created_at).total_seconds() <= 3600

    # Registration and its one-time rewards (first-contest coins, early-bird
    # badge) must commit together so a failure can't leave a registration
    # whose rewards were already paid out, or vice versa.
    with transaction.atomic():
        reg, created = ContestRegistration.objects.update_or_create(
            contest=contest, user=request.user,
            defaults={'is_rated': want_rated, 'is_early_bird': is_early and is_first},
        )

        if created:
            if is_first and ContestRegistration.objects.filter(user=request.user).count() == 1:
                coin_svc.award_coins(request.user, 'first_contest', contest=contest,
                                     note='Your very first contest!')
            if is_early:
                badge_svc.award_early_bird(request.user, contest)

    if created:
        messages.success(request, f'Registered for {contest.title} '
                                  f'({"rated" if want_rated else "unrated"}).')
    else:
        messages.info(request, f'Updated registration to '
                               f'{"rated" if want_rated else "unrated"}.')

    return redirect('contest_detail', pk=pk)


@login_required
def contest_set_rated(request, pk):
    """POST: toggle rated/unrated participation before contest start."""
    from .models import Contest, ContestRegistration
    from django.utils import timezone
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST only'}, status=405)
    contest = get_object_or_404(Contest, pk=pk)
    if timezone.now() >= contest.start_time:
        return JsonResponse({'ok': False, 'error': 'Contest has started'}, status=400)
    reg, _ = ContestRegistration.objects.get_or_create(
        contest=contest, user=request.user,
    )
    want = request.POST.get('is_rated', '1') == '1'
    if not contest.is_rated:
        want = False
    if not contest.allow_unrated_join and contest.is_rated:
        want = True
    reg.is_rated = want
    reg.save(update_fields=['is_rated'])
    return JsonResponse({'ok': True, 'is_rated': reg.is_rated})


@login_required
def leaderboard_data(request, pk):
    """JSON leaderboard for AJAX polling."""
    from .models import Contest, ContestSubmission, UserRating
    from django.http import JsonResponse
    from django.utils import timezone

    contest = get_object_or_404(Contest, pk=pk)
    now = timezone.now()
    if contest.hide_leaderboard_until_end and contest.start_time <= now <= contest.end_time:
        return JsonResponse({
            'hidden': True,
            'rows': [],
            'updated_at': now.isoformat(),
        })

    subs = ContestSubmission.objects.filter(
        contest=contest, is_submitted=True,
        student__profile__role='STUDENT',
    ).select_related('student').order_by('-total_marks', 'duration_taken')

    ratings = {
        ur.user_id: ur for ur in UserRating.objects.filter(
            user_id__in=[s.student_id for s in subs]
        )
    }
    rows = []
    for idx, s in enumerate(subs):
        ur = ratings.get(s.student_id)
        title = ur.rank_title if ur else {'title': 'Newcomer', 'color': '#808080'}
        rows.append({
            'rank': idx + 1,
            'username': s.student.username,
            'score': s.total_marks,
            'time_taken': s.duration_taken,
            'rating': ur.rating if ur else 1000,
            'rank_title': title['title'],
            'rank_color': title['color'],
            'rating_change': s.rating_change,
            'percentile': s.percentile,
            'is_virtual': s.is_virtual,
            'is_me': s.student_id == request.user.id,
        })
    return JsonResponse({
        'hidden': False,
        'rows': rows,
        'updated_at': now.isoformat(),
        'contest_status': 'live' if contest.start_time <= now <= contest.end_time
                          else ('past' if now > contest.end_time else 'upcoming'),
    })


@login_required
def virtual_contest(request, pk):
    """Replay a past contest as a virtual practice attempt (no rating change)."""
    from .models import Contest, VirtualContest, ContestSubmission
    from django.utils import timezone

    contest = get_object_or_404(Contest, pk=pk)
    if not contest.allows_virtual:
        messages.error(request, 'Virtual replay is disabled for this contest.')
        return redirect('contest_detail', pk=pk)
    if timezone.now() < contest.end_time:
        messages.error(request, 'Virtual contests are only available after the contest ends.')
        return redirect('contest_detail', pk=pk)

    vc, _ = VirtualContest.objects.get_or_create(
        user=request.user, contest=contest, finished_at__isnull=True,
        defaults={},
    )

    sub, _ = ContestSubmission.objects.get_or_create(
        contest=contest, student=request.user,
        defaults={'is_virtual': True, 'is_rated_participant': False},
    )
    if not sub.is_virtual:
        sub.is_virtual = True
        sub.is_rated_participant = False
        sub.save(update_fields=['is_virtual', 'is_rated_participant'])

    questions = contest.questions.all()
    return render(request, 'core/virtual_contest.html', {
        'contest': contest,
        'questions': questions,
        'submission': sub,
        'virtual': vc,
    })


@login_required
def badge_gallery(request):
    """Public badge gallery: shows all badges + which the user has earned."""
    from .models import Badge, UserBadge
    badges = Badge.objects.filter(is_active=True).order_by('badge_type', 'rarity', 'name')
    earned = {
        ub.badge_id: ub for ub in UserBadge.objects.filter(user=request.user).select_related('badge')
    }
    grouped = {}
    for b in badges:
        grouped.setdefault(b.get_badge_type_display(), []).append({
            'badge': b,
            'earned': b.id in earned,
            'earned_at': earned[b.id].earned_at if b.id in earned else None,
        })
    total_badges = badges.count()
    earned_count = len(earned)
    rarest = None
    rarity_order = {'legendary': 4, 'epic': 3, 'rare': 2, 'common': 1}
    for b in badges:
        if b.id in earned:
            if rarest is None or rarity_order[b.rarity] > rarity_order[rarest.rarity]:
                rarest = b
    return render(request, 'core/badge_gallery.html', {
        'grouped': grouped,
        'total_badges': total_badges,
        'earned_count': earned_count,
        'rarest': rarest,
    })


@login_required
def profile_contests(request):
    """Codeforces-style contest profile: rating chart, badges, history."""
    from .models import (
        UserRating, ContestRatingHistory, UserBadge, Badge, ContestSubmission,
    )
    from django.db.models import Avg
    from django.core.paginator import Paginator
    from django.utils import timezone

    rating, _ = UserRating.objects.get_or_create(user=request.user)
    history = ContestRatingHistory.objects.filter(
        user=request.user,
    ).select_related('contest').order_by('recorded_at')[:20]

    earned = UserBadge.objects.filter(
        user=request.user,
    ).select_related('badge').order_by('-earned_at')
    earned_ids = set(eb.badge_id for eb in earned)
    all_badges = Badge.objects.filter(is_active=True).order_by('badge_type', 'rarity')

    badge_list = []
    for b in all_badges:
        badge_list.append({
            'badge': b,
            'earned': b.id in earned_ids,
        })

    submissions = ContestSubmission.objects.filter(
        student=request.user, is_submitted=True, is_virtual=False,
    ).select_related('contest').order_by('-submitted_at')
    paginator = Paginator(submissions, 15)
    page = paginator.get_page(request.GET.get('page'))

    stats = {
        'entered': rating.contests_entered,
        'best_rank': rating.best_rank or '—',
        'wins': submissions.filter(rank_in_contest=1).count(),
        'avg_percentile': submissions.aggregate(a=Avg('percentile'))['a'] or 0,
    }
    if rating.contests_entered:
        stats['win_rate'] = round((stats['wins'] / rating.contests_entered) * 100, 1)
    else:
        stats['win_rate'] = 0
    stats['avg_percentile'] = round(stats['avg_percentile'], 1)

    chart_data = [{
        'contest': h.contest.title[:30],
        'rating': h.new_rating,
        'change': h.change,
        'date': h.recorded_at.strftime('%Y-%m-%d'),
    } for h in history]

    # Activity calendar (last 52 weeks)
    from datetime import timedelta
    today = timezone.now().date()
    days = []
    activity_map = {}
    for s in submissions:
        if not s.submitted_at:
            continue
        d = s.submitted_at.date()
        score = 1
        if s.percentile is not None:
            if s.percentile <= 10 or s.rank_in_contest == 1:
                score = 4
            elif s.percentile <= 25:
                score = 3
            elif s.percentile <= 50:
                score = 2
        prev = activity_map.get(d, (0, None))
        if score > prev[0]:
            activity_map[d] = (score, s.contest.title)
    start = today - timedelta(days=7 * 52)
    cur = start
    while cur <= today:
        level, label = activity_map.get(cur, (0, None))
        days.append({
            'date': cur.isoformat(),
            'level': level,
            'label': label,
        })
        cur += timedelta(days=1)

    return render(request, 'core/profile_contests.html', {
        'rating': rating,
        'history': history,
        'chart_data_json': json.dumps(chart_data),
        'earned_badges': earned,
        'badge_list': badge_list,
        'submissions_page': page,
        'paginator': paginator,
        'stats': stats,
        'activity_days': days,
    })


@login_required
def coin_balance_api(request):
    from .models import UserRating
    from django.http import JsonResponse
    rating, _ = UserRating.objects.get_or_create(user=request.user)
    return JsonResponse({
        'balance': rating.coin_balance,
        'rating': rating.rating,
        'rank_title': rating.rank_title['title'],
    })


@login_required
def check_badges_api(request):
    """POST: run the badge engine and return any newly awarded badges."""
    from .services import badges as badge_svc
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    new_badges = badge_svc.check_and_award_badges(request.user)
    return JsonResponse({
        'ok': True,
        'new_badges': [
            {
                'name': b.name,
                'icon': b.icon,
                'rarity': b.rarity,
                'color_hex': b.color_hex,
                'description': b.description,
            }
            for b in new_badges
        ],
    })
